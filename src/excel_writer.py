# -*- coding: utf-8 -*-
"""
按配置将内容写入 Excel 的指定工作表与单元格位置。
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def ensure_sheet(wb: Workbook, sheet_name: str) -> Worksheet:
    """
    若工作表不存在则创建。
    
    Args:
        wb: 工作簿对象
        sheet_name: 工作表名称
        
    Returns:
        工作表对象
    """
    if not sheet_name or not sheet_name.strip():
        raise ValueError("工作表名称不能为空")
    
    sheet_name = sheet_name.strip()
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    return wb[sheet_name]


def cell_to_row_col(cell: str) -> tuple[int, int]:
    """
    将 Excel 单元格地址如 'B3' 转为 (row, col) 从 1 开始。
    
    Args:
        cell: Excel 单元格地址（如 'B3', 'AA10'）
        
    Returns:
        (行号, 列号) 元组，从 1 开始
        
    Raises:
        ValueError: 单元格地址格式无效
    """
    if not cell or not isinstance(cell, str):
        raise ValueError(f"无效的单元格地址: {cell}")
    
    cell = cell.strip().upper()
    col_part = ""
    row_part = ""
    for c in cell:
        if c.isalpha():
            col_part += c
        elif c.isdigit():
            row_part += c
        else:
            raise ValueError(f"单元格地址包含无效字符: {cell}")
    
    if not col_part:
        raise ValueError(f"单元格地址缺少列标识: {cell}")
    
    row = int(row_part) if row_part else 1
    if row < 1:
        raise ValueError(f"行号必须 >= 1: {cell}")
    
    col = 0
    for c in col_part:
        col = col * 26 + (ord(c) - ord("A") + 1)
    
    if col < 1 or col > 16384:  # Excel 最大列数
        raise ValueError(f"列号超出范围 (1-16384): {cell}")
    
    return row, col


def write_cell(ws: Worksheet, cell_address: str, value: Any) -> None:
    """
    向指定单元格写入一个值。
    
    Args:
        ws: 工作表对象
        cell_address: 单元格地址（如 'B3'）
        value: 要写入的值
    """
    try:
        r, c = cell_to_row_col(cell_address)
        ws.cell(row=r, column=c, value=value)
    except Exception as e:
        logger.error(f"写入单元格 {cell_address} 失败: {e}")
        raise


def write_table(ws: Worksheet, start_cell: str, table: list[list[Any]]) -> None:
    """
    从 start_cell 开始按行、列写入二维表格。
    
    Args:
        ws: 工作表对象
        start_cell: 起始单元格地址（如 'B3'）
        table: 行→列 的二维列表
    """
    if not table:
        logger.warning("表格为空，跳过写入")
        return
    
    try:
        start_row, start_col = cell_to_row_col(start_cell)
        for ri, row in enumerate(table):
            if not isinstance(row, (list, tuple)):
                logger.warning(f"第 {ri} 行不是列表或元组，跳过")
                continue
            for ci, val in enumerate(row):
                try:
                    ws.cell(row=start_row + ri, column=start_col + ci, value=val)
                except Exception as e:
                    logger.error(f"写入单元格 ({start_row + ri}, {start_col + ci}) 失败: {e}")
    except Exception as e:
        logger.error(f"写入表格失败: {e}")
        raise


def write_text_block(ws: Worksheet, start_cell: str, text: str, max_chars_per_cell: int = 32000) -> None:
    """
    将长文本写入从 start_cell 开始的一个单元格（若超长则截断，Excel 单格约 32k 字符）。
    
    Args:
        ws: 工作表对象
        start_cell: 起始单元格地址
        text: 要写入的文本
        max_chars_per_cell: 单个单元格最大字符数（默认 32000）
    """
    if not isinstance(text, str):
        text = str(text) if text is not None else ""
    
    if max_chars_per_cell < 1:
        max_chars_per_cell = 32000
        logger.warning(f"max_chars_per_cell 无效，重置为 {max_chars_per_cell}")
    
    if len(text) > max_chars_per_cell:
        text = text[: max_chars_per_cell - 3] + "..."
        logger.warning(f"文本被截断到 {max_chars_per_cell} 字符")
    
    write_cell(ws, start_cell, text)


def load_or_create_workbook(excel_path: str | Path) -> Workbook:
    """
    若文件存在则加载，否则创建新工作簿。
    
    Args:
        excel_path: Excel 文件路径
        
    Returns:
        工作簿对象
        
    Raises:
        ValueError: 文件路径无效或文件损坏
    """
    excel_path = Path(excel_path)
    
    if excel_path.exists():
        if not excel_path.is_file():
            raise ValueError(f"路径不是文件: {excel_path}")
        try:
            return load_workbook(excel_path)
        except Exception as e:
            raise ValueError(f"无法加载 Excel 文件 {excel_path}: {e}") from e
    
    return Workbook()


def save_workbook(wb: Workbook, excel_path: str | Path) -> None:
    """
    保存工作簿。
    
    Args:
        wb: 工作簿对象
        excel_path: 保存路径
        
    Raises:
        ValueError: 保存失败
    """
    excel_path = Path(excel_path)
    try:
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(excel_path)
        logger.info(f"工作簿已保存: {excel_path}")
    except Exception as e:
        logger.error(f"保存工作簿失败: {e}")
        raise ValueError(f"无法保存 Excel 文件 {excel_path}: {e}") from e
