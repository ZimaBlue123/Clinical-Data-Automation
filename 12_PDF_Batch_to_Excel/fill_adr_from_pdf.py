"""
从 不同剂量组ADR分析 (TFL).pdf 与 不同剂量组ADR分析-分级 (TFL).pdf 提取数据，
按“发热”的填写逻辑，填充 不同剂量组ADR分析 (TFL).xlsx 中所有 ADR 的 1级/2级/3级/Total。
Excel 列顺序：低剂量试验组(例数/人数/发生率)、高剂量试验组、低剂量佐剂组、高剂量佐剂组、安慰剂组。
"""

import re
from pathlib import Path
from openpyxl import load_workbook
from src.pdf_reader import extract_tables_from_pdf
import sys

BASE = Path(__file__).resolve().parent
PDF_TFL = BASE / "不同剂量组ADR分析 (TFL).pdf"
PDF_GRADE = BASE / "不同剂量组ADR分析-分级 (TFL).pdf"
XLSX_PATH = BASE / "不同剂量组ADR分析 (TFL).xlsx"

# PDF 表列顺序：首选术语, 低剂量佐剂(n%,例次), 高剂量佐剂, 低剂量试验, 高剂量试验, 安慰剂, 合计, P值
# 即 0=术语, 1-2=低剂量佐, 3-4=高剂量佐, 5-6=低剂量试, 7-8=高剂量试, 9-10=安慰剂
# Excel 列顺序：低剂量试验, 高剂量试验, 低剂量佐剂, 高剂量佐剂, 安慰剂
PDF_GROUP_ORDER = [
    (5, 6),  # 低剂量试验组 -> Excel 第1组
    (7, 8),  # 高剂量试验组 -> Excel 第2组
    (1, 2),  # 低剂量佐剂组 -> Excel 第3组
    (3, 4),  # 高剂量佐剂组 -> Excel 第4组
    (9, 10),  # 安慰剂组 -> Excel 第5组
]

# Excel 中 ADR 名称与 PDF 首选术语的映射（Excel 用简称时）
EXCEL_TO_PDF_TERM = {
    "发热": "发热",
    "注射部位疼痛": "疫苗接种部位疼痛",
    "疫苗接种部位疼痛": "疫苗接种部位疼痛",
    "疲劳": "疲劳",
    "疼痛": "肌痛",
    "肌痛": "肌痛",
    "注射部位肿胀": "疫苗接种部位肿胀",
    "疫苗接种部位肿胀": "疫苗接种部位肿胀",
    "头痛": "头痛",
    "注射部位红肿": "疫苗接种部位红斑",
    "疫苗接种部位红斑": "疫苗接种部位红斑",
    "关节痛": "关节痛",
    "注射部位瘙痒": "疫苗接种部位瘙痒",
    "疫苗接种部位瘙痒": "疫苗接种部位瘙痒",
    "注射部位硬结": "疫苗接种部位硬结",
    "疫苗接种部位硬结": "疫苗接种部位硬结",
    "面痛": "面痛",
    "头晕": "头晕",
    "呼吸困难": "呼吸困难",
    "咳嗽": "咳嗽",
    "口咽疼痛": "口咽疼痛",
    "流涕": "流涕",
    "恶心": "恶心",
    "室上性期外收缩": "室上性期外收缩",
    "尿蛋白检出": "尿蛋白检出",
    "丙氨酸氨基转移酶升高": "丙氨酸氨基转移酶升高",
    "尿红细胞阳性": "尿红细胞阳性",
    "天门冬氨酸氨基转移酶升高": "天门冬氨酸氨基转移酶升高",
    "心电图T波异常": "心电图T波异常",
    "血葡萄糖升高": "血葡萄糖升高",
    "白细胞计数升高": "白细胞计数升高",
    "淋巴细胞百分比升高": "淋巴细胞百分比升高",
    "淋巴细胞计数降低": "淋巴细胞计数降低",
    "尿糖检出": "尿糖检出",
    "心电图ST段异常": "心电图ST段异常",
    "血纤维蛋白原升高": "血纤维蛋白原升高",
    "中性粒细胞计数升高": "中性粒细胞计数升高",
}


def parse_n_pct(s: str):
    """从 '3 ( 30.00)' 解析出 (人数, 发生率)。"""
    if not s or not isinstance(s, str):
        return None, None
    m = re.match(r"\s*(\d+)\s*\(\s*[\d.]+\s*\)", s.strip())
    if not m:
        return None, None
    n = int(m.group(1))
    pct_m = re.search(r"\(\s*([\d.]+)\s*\)", s)
    rate = float(pct_m.group(1)) / 100.0 if pct_m else None
    return n, rate


def parse_example_count(s: str):
    """例次：取整数。"""
    if s is None:
        return None
    s = str(s).strip()
    if not s or s in ("-", ""):
        return 0
    try:
        return int(float(s))
    except ValueError:
        return None


def parse_row_to_groups(row: list, group_order=None):
    """将 PDF 表的一行解析为 5 组的 (例次, 人数, 发生率)。"""
    group_order = group_order or PDF_GROUP_ORDER
    result = []
    for idx_n, idx_ex in group_order:
        while len(row) <= max(idx_n, idx_ex):
            row.append("")
        n_pct = (row[idx_n] or "").strip() if idx_n < len(row) else ""
        ex = row[idx_ex] if idx_ex < len(row) else ""
        n_subj, rate = parse_n_pct(n_pct)
        ex_count = parse_example_count(ex)
        if n_subj is None and ex_count is None:
            result.append((None, None, None))
        else:
            result.append(
                (
                    ex_count if ex_count is not None else 0,
                    n_subj if n_subj is not None else 0,
                    rate if rate is not None else 0.0,
                )
            )  # noqa: E501
    return result


def first_nonempty_cell(row: list, max_col: int = 4) -> str:
    """跨页时首列可能为空，取前 max_col 列中第一个非空字符串。"""
    if not row:
        return ""
    for c in range(min(max_col, len(row))):
        v = row[c]
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def is_grade_row(first_cell: str):
    """是否为分级行：1级、2级、3级、4级、5级、≥3级。"""
    if not first_cell:
        return False
    s = str(first_cell).strip()
    return s in ("1级", "2级", "3级", "4级", "5级", "≥3级")


def is_total_row(first_cell: str):
    """是否为 Total 行（分级表中有时用 Total）。"""
    return str(first_cell or "").strip() == "Total"


def is_adr_or_soc(first_cell: str):
    """可能是首选术语或系统器官（ADR 名）。"""
    if not first_cell or not str(first_cell).strip():
        return False
    s = str(first_cell).strip()
    if s in ("首选术语", "系统器官", "年龄组", "不良事件", "n(%)", "例次"):
        return False
    if s in ("1级", "2级", "3级", "4级", "5级", "≥3级", "Total"):
        return False
    return not re.match(r"^[\d\s.()%]+$", s)


def is_likely_soc(first_cell: str) -> bool:
    """系统器官类（SOC）行有数字但不应作为“当前 ADR”，否则会抢掉后续 1/2/3 级。"""
    if not first_cell:
        return False
    s = str(first_cell).strip()
    if "系统" in s or "疾病" in s or s.endswith("反应") or s.endswith("检查"):
        return True
    return bool("及" in s and ("组织" in s or "器官" in s))


def is_footer_or_header(first_cell: str) -> bool:
    """表头/页脚行不作为 ADR，不更新 last_pt。"""
    if not first_cell:
        return True
    fc = str(first_cell).strip()
    skip = (
        "使用MedDRA",
        "数据来源",
        "百分比",
        "合同研究",
        "方案编号",
        "年龄组",
        "系统器官",
        "首选术语",
        "低剂量佐剂组",
        "高剂量佐剂组",
        "低剂量试验组",
        "高剂量试验组",
        "安慰剂组",
        "合计",
        "若受试者",
        "远大赛威信",
    )
    return any(fc.startswith(s) or s in fc for s in skip)


def collect_graded_adr_data(pdf_path: Path):  # noqa: PLR0915 - TODO: 下个迭代重构 # noqa: PLR0912 - TODO: 下个迭代重构
    """
    从分级 PDF 中收集：每个 ADR 的 Total 行 + 1级/2级/3级 行对应的 5 组数据。
    分级表可能跨页，1级/2级/3级 在下一页，故用“当前 ADR”关联后续遇到的 1级/2级/3级。
    返回: { "发热": { "Total": [(例数,人数,发生率)*5], "1级": [...], "2级": [...], "3级": [...] }, ... }
    """
    tables_by_page = extract_tables_from_pdf(pdf_path)
    all_rows = []
    for page_tables in tables_by_page:
        for table in page_tables:
            if not table:
                continue
            for row in table:
                if not row or not any(c for c in row):
                    continue
                all_rows.append([str(c).strip() if c else "" for c in row])

    adr_data = {}
    last_pt = None  # 仅首选术语（PT），用于关联后续 1/2/3 级；SOC 不更新
    i = 0
    while i < len(all_rows):
        row = all_rows[i]
        first = first_nonempty_cell(row)
        if not first:
            i += 1
            continue
        if is_adr_or_soc(first) and not is_grade_row(first):
            parsed = parse_row_to_groups(row)
            has_data = any(p[0] is not None or p[1] is not None for p in parsed)
            if has_data and not is_likely_soc(first) and not is_footer_or_header(first):
                last_pt = first
                if last_pt not in adr_data:
                    adr_data[last_pt] = {}
                adr_data[last_pt]["Total"] = parsed
            i += 1
            continue
        if first == "1级" and last_pt is not None:
            # 避免把 SOC（如“各类检查”）下面的 1/2/3 级误归到上一个首选术语并覆盖其分级
            prev_label = ""
            back = i - 1
            while back >= 0:
                bfc = first_nonempty_cell(all_rows[back])
                if not bfc or is_footer_or_header(bfc) or is_grade_row(bfc):
                    back -= 1
                    continue
                prev_label = bfc
                break
            if prev_label and is_likely_soc(prev_label):
                i += 1
                continue

            if last_pt not in adr_data:
                adr_data[last_pt] = {}
            adr_data[last_pt]["1级"] = parse_row_to_groups(row)
            i += 1

            # 跨页时 2级/3级 可能不在紧邻的下一行（中间有 4级、5级、≥3级 或换页表头），向前扫描
            # 仅当遇到“带数字的另一条首选术语行”时才停止；表头/页脚不打断
            def is_new_adr_row(r):
                fc = first_nonempty_cell(r)
                if is_footer_or_header(fc) or is_grade_row(fc):
                    return False
                if not is_adr_or_soc(fc):
                    return False
                p = parse_row_to_groups(r)
                return any(x[0] is not None or x[1] is not None for x in p)

            j = i
            while j < len(all_rows):
                fc = first_nonempty_cell(all_rows[j])
                if fc == "2级":
                    adr_data[last_pt]["2级"] = parse_row_to_groups(all_rows[j])
                    j += 1
                    break
                if is_new_adr_row(all_rows[j]):
                    break
                j += 1
            while j < len(all_rows):
                fc = first_nonempty_cell(all_rows[j])
                if fc == "3级":
                    adr_data[last_pt]["3级"] = parse_row_to_groups(all_rows[j])
                    j += 1
                    break
                if is_new_adr_row(all_rows[j]):
                    break
                j += 1
            i = j
            continue
        i += 1

    return adr_data


def find_pdf_term_for_excel(excel_adr: str) -> str:
    """根据 Excel 中的 ADR 名称找到 PDF 中的首选术语。"""
    excel_adr = (excel_adr or "").strip()
    if not excel_adr:
        return ""
    return EXCEL_TO_PDF_TERM.get(excel_adr, excel_adr)


def fill_excel():  # noqa: PLR0915 - TODO: 下个迭代重构 # noqa: PLR0912 - TODO: 下个迭代重构
    """
    从分级 PDF 提取数据并填充到 Excel。

    Returns:
        填充的行数

    Raises:
        FileNotFoundError: PDF 或 Excel 文件不存在
        ValueError: 数据格式错误
    """
    if not PDF_GRADE.exists():
        raise FileNotFoundError(f"分级 PDF 不存在: {PDF_GRADE}")

    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {XLSX_PATH}")

    try:
        adr_data = collect_graded_adr_data(PDF_GRADE)
    except Exception as e:
        raise ValueError(f"解析分级 PDF 失败: {e}") from e

    # 若分级 PDF 中某 ADR 只有 Total 没有 1/2/3 级，可从非分级 PDF 取 Total，分级仍用分级 PDF
    try:
        wb = load_workbook(XLSX_PATH)
    except Exception as e:
        raise ValueError(f"无法加载 Excel 文件: {e}") from e

    try:
        ws = wb.active
        # 确定数据列范围：C 到 Q（5 组 × 3）
        start_col = 3
        cells_per_group = 3  # 例数, 人数, 发生率
        grade_rows = ("1级", "2级", "3级", "Total")
        row = 1
        filled = 0

        while row <= ws.max_row:
            try:
                grade_cell = ws.cell(row=row, column=2).value
                grade = (grade_cell or "").strip() if grade_cell else ""
                if grade != "1级":
                    row += 1
                    continue

                # 向上找 ADR 名（A 列，仅第一行有）
                a = row
                while a >= 1:
                    ac = ws.cell(row=a, column=1).value
                    if ac is not None and str(ac).strip():
                        break
                    a -= 1

                if a < 1:
                    row += 1
                    continue

                adr_name = (ws.cell(row=a, column=1).value or "").strip()
                pdf_term = find_pdf_term_for_excel(adr_name)

                if not pdf_term or pdf_term not in adr_data:
                    row += 1
                    continue

                data = adr_data[pdf_term]
                for ri, gr in enumerate(grade_rows):
                    r = row + ri
                    if r > ws.max_row:
                        break
                    if gr not in data:
                        continue

                    vals = data[gr]
                    if len(vals) != 5:
                        continue  # 跳过格式不正确的数据

                    for gi, val_tuple in enumerate(vals):
                        if gi >= 5:  # 只处理5组
                            break
                        if not isinstance(val_tuple, (tuple, list)) or len(val_tuple) < 3:
                            continue

                        ex_count, n_subj, rate = val_tuple[0], val_tuple[1], val_tuple[2]
                        c = start_col + gi * cells_per_group
                        # Excel 列为：例数(=人数 n)、例次、发生率
                        if n_subj is not None:
                            ws.cell(row=r, column=c, value=n_subj)
                        if ex_count is not None:
                            ws.cell(row=r, column=c + 1, value=ex_count)
                        if rate is not None:
                            ws.cell(row=r, column=c + 2, value=rate)
                    filled += 1

                row += 4
            except Exception as e:
                print(f"处理第 {row} 行时出错: {e}")
                row += 1
                continue

        wb.save(XLSX_PATH)
        print(f"已填充 {filled} 行数据，已保存: {XLSX_PATH}")
        return filled

    except Exception as e:
        raise ValueError(f"填充 Excel 数据失败: {e}") from e
    finally:
        wb.close()


if __name__ == "__main__":
    try:
        fill_excel()
    except FileNotFoundError as e:
        print(f"文件未找到: {e}")
        sys.exit(1)
    except ValueError as e:
        print(f"数据错误: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"未知错误: {e}")
        import traceback

        traceback.print_exc()
        sys.exit(1)
