"""
中检院血清样本检测报告 PDF 转 Excel。

- 支持矢量表（pdfplumber.extract_tables）：按表头识别五项；每项输出「检测值」+「说明」两列。
- 扫描件：可选 Tesseract OCR（chi_sim+eng）按行解析。

Excel 版式与报告一致：样品 ID + 五项（每项 2 列：数值、说明），列顺序为
Anti-HBs → HBsAg → Anti-HBc → Anti-HBe → HBeAg；缺项留空。
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from collections import OrderedDict
from pathlib import Path
from typing import Any
from collections.abc import Iterable

# 项目根：12_PDF_Batch_to_Excel -> Clinical Data Automation
MODULE_DIR = Path(__file__).resolve().parent
ROOT = MODULE_DIR.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from src.pdf_reader import extract_tables_from_pdf  # noqa: E402 — 须在 sys.path.insert 之后
from src.serology_utils import OUTPUT_MARKERS, MARKER_UNITS, canonical_sample_id  # noqa: E402

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None  # type: ignore

try:
    import pytesseract
    from PIL import Image

    _OCR_AVAILABLE = True
except ImportError:
    _OCR_AVAILABLE = False

logger = logging.getLogger(__name__)

# 表头 / OCR 中可能出现的写法 -> 规范名
MARKER_ALIASES: dict[str, str] = {
    "anti-hbe": "Anti-HBe",
    "hbeag": "HBeAg",
    "hbsag": "HBsAg",
    "anti-hbc": "Anti-HBc",
    "anti-hbs": "Anti-HBs",
}


def _norm_cell(x: Any) -> str:
    if x is None:
        return ""
    return str(x).strip()


def _canonical_marker(cell: str) -> str | None:
    s = _norm_cell(cell).lower().replace(" ", "")
    if not s:
        return None
    for key, canon in MARKER_ALIASES.items():
        if key in s or s == key.replace("-", ""):
            return canon
    # 中英混排时保留原关键片段
    for canon in OUTPUT_MARKERS:
        if canon.lower() in s or canon.replace("-", "").lower() in s:
            return canon
    return None


def _has_sample_id_header(row: list[Any]) -> bool:
    t = " ".join(_norm_cell(c) for c in row[:4])
    return "样品" in t and "ID" in t


def _marker_column_starts(header_row: list[Any]) -> dict[int, str]:
    """列索引 -> 规范指标名（每个指标占连续多列：数值 + 说明等）。"""
    out: dict[int, str] = {}
    for j, cell in enumerate(header_row):
        m = _canonical_marker(_norm_cell(cell))
        if m:
            out[j] = m
    return out


def _build_column_groups(header_row: list[Any], sub_row: list[Any] | None) -> list[tuple[str, int, int]]:
    """
    返回 [(指标名, 起始列, 结束列开区间), ...] 按从左到右顺序。
    结束列为下一指标起始或表宽。
    """
    starts = sorted(_marker_column_starts(header_row).items())
    if not starts:
        return []
    width = max(len(header_row), len(sub_row or []))
    groups: list[tuple[str, int, int]] = []
    for i, (col, name) in enumerate(starts):
        end = starts[i + 1][0] if i + 1 < len(starts) else width
        groups.append((name, col, end))
    return groups


def _is_note_cell(text: str) -> bool:
    t = _norm_cell(text)
    if not t:
        return False
    if t in ("阴性", "阳性", "反应", "可疑", "待查"):
        return True
    return bool(len(t) <= 4 and re.match(r"^[\u4e00-\u9fff]+$", t))


def _extract_group_value_and_note(  # noqa: PLR0912 - TODO: 下个迭代重构
    row: list[str],
    c0: int,
    c1: int,
    sub_row: list[str] | None,
) -> tuple[str, str]:
    """从 row[c0:c1] 中取（检测值, 说明）；优先按子表头「说明」列对齐。"""
    if sub_row and len(sub_row) >= c0:
        note_j: int | None = None
        for j in range(c0, min(c1, len(sub_row))):
            if _norm_cell(sub_row[j]) == "说明":
                note_j = j
                break
        if note_j is not None:
            note = _norm_cell(row[note_j]) if note_j < len(row) else ""
            val_parts: list[str] = []
            for k in range(c0, note_j):
                if k >= len(row):
                    continue
                sub_k = _norm_cell(sub_row[k]) if k < len(sub_row) else ""
                if sub_k == "说明":
                    continue
                v = _norm_cell(row[k])
                if v:
                    if re.search(r"[\d<≥]", v) or re.match(r"^[\d.<>≥]+$", v):
                        val_parts.append(v.split()[0] if v.split() else v)
                    elif not _is_note_cell(v):
                        val_parts.append(v)
            val = " ".join(val_parts).strip()
            if not val and note_j > c0:
                val = _norm_cell(row[note_j - 1]) if note_j - 1 >= c0 else ""
            return val, note

    # 无子表头：最后一格像说明则拆开，否则整段当数值
    cells = [_norm_cell(row[i]) if i < len(row) else "" for i in range(c0, min(c1, len(row)))]
    vals_only = [c for c in cells if c]
    if not vals_only:
        return "", ""
    if len(vals_only) >= 2 and _is_note_cell(vals_only[-1]):
        note = vals_only[-1]
        rest = vals_only[:-1]
        num_like = []
        for x in rest:
            if re.search(r"[\d<≥]", x) or re.match(r"^[\d.<>≥]+$", x):
                num_like.append(x.split()[0] if x.split() else x)
        val = " ".join(num_like).strip() if num_like else " ".join(rest).strip()
        return val, note
    num_like = []
    for c in vals_only:
        if re.search(r"[\d<≥]", c) or re.match(r"^[\d.<>≥]+$", c):
            num_like.append(c.split()[0] if c.split() else c)
    if num_like:
        return " ".join(num_like).strip(), ""
    return vals_only[0] if vals_only else "", ""


def _iter_body_rows(table: list[list[Any]], start_i: int) -> Iterable[list[str]]:
    for r in range(start_i, len(table)):
        row = [_norm_cell(table[r][j]) if j < len(table[r]) else "" for j in range(len(table[r]))]
        if not any(row):
            continue
        yield row


def parse_digital_table(table: list[list[Any]]) -> list[dict[str, Any]]:  # noqa: PLR0912 - TODO: 下个迭代重构
    if not table or len(table) < 2:
        return []
    header_idx = None
    for hi, row in enumerate(table[:6]):
        if _marker_column_starts(row) and (_has_sample_id_header(row) or len(_marker_column_starts(row)) >= 1):
            header_idx = hi
            break
    if header_idx is None:
        return []

    header = table[header_idx]
    sub = table[header_idx + 1] if header_idx + 1 < len(table) else None
    groups = _build_column_groups(header, sub)
    if not groups:
        return []

    data_start = header_idx + 1
    if sub and any(_norm_cell(c) in ("mIU/ml", "IU/ml", "S/CO", "说明") for c in sub):
        data_start = header_idx + 2

    rows_out: list[dict[str, Any]] = []
    for row in _iter_body_rows(table, data_start):
        if len(row) < 2:
            continue
        sid = row[1] if row[1] else row[0]
        if not sid or sid in ("样品", "ID", "序号"):
            continue
        if re.match(r"^[\d.]+$", sid) and not re.search(r"-", sid):
            continue
        w = max(len(row), len(sub or []), len(header))
        sub_row = [(_norm_cell(sub[j]) if sub and j < len(sub) else "") for j in range(w)]
        rec: dict[str, Any] = {"样品ID": sid}
        for m in OUTPUT_MARKERS:
            rec[m] = {"value": "", "note": ""}
        for name, c0, c1 in groups:
            if name in OUTPUT_MARKERS:
                v, n = _extract_group_value_and_note(row, c0, c1, sub_row)
                rec[name] = {"value": v, "note": n}
        rows_out.append(rec)
    return rows_out


def extract_rows_digital_pdf(pdf_path: Path) -> list[dict[str, Any]]:
    all_rows: list[dict[str, Any]] = []
    try:
        tables_by_page = extract_tables_from_pdf(pdf_path)
    except Exception as e:
        logger.warning("pdfplumber 读表失败: file=%s reason=%s", pdf_path.name, e)
        return []
    for page_tables in tables_by_page:
        for table in page_tables:
            all_rows.extend(parse_digital_table(table))
    return all_rows


def _clean_ocr_value(raw: str) -> str:
    s = _norm_cell(raw).replace("��", "").strip()
    if not s:
        return ""
    m = re.match(r"^([<≥\d.]+)\s*", s)
    if m:
        return m.group(1).strip()
    return s.split()[0] if s.split() else s


def _clean_ocr_note(raw: str) -> str:
    return _norm_cell(raw.replace("��", ""))


def _normalize_sample_id(raw: str) -> str:
    s = _norm_cell(raw)
    if not s:
        return ""
    s = s.replace("—", "-").replace("–", "-").replace("_", "-")
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    parts = s.split("-")
    if len(parts) >= 3:
        p0, p1, p2 = parts[0], parts[1], parts[2]
        # 常见 OCR 误识别：M1 -> Ml / MI
        p1 = re.sub(r"^M[Il]$", "M1", p1, flags=re.IGNORECASE)
        p1 = re.sub(r"^M([Il])(\d+)$", r"M1\2", p1, flags=re.IGNORECASE)
        # 首段通常是序号，允许把 O 纠正为 0
        if re.search(r"\d", p0):
            p0 = p0.replace("O", "0").replace("o", "0")
        return f"{p0}-{p1}-{p2}"
    return s


def _is_empty_marker(block: dict[str, Any]) -> bool:
    return not ((block.get("value") or "").strip() or (block.get("note") or "").strip())


def _load_reference_excel(path: Path) -> dict[str, dict[str, dict[str, str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out: dict[str, dict[str, dict[str, str]]] = {}
    for r in range(3, ws.max_row + 1):
        sid_raw = ws.cell(r, 1).value
        sid = canonical_sample_id(str(sid_raw) if sid_raw is not None else "")
        if not sid:
            continue
        rec: dict[str, dict[str, str]] = {}
        col = 2
        for m in OUTPUT_MARKERS:
            v = ws.cell(r, col).value
            n = ws.cell(r, col + 1).value
            rec[m] = {
                "value": str(v).strip() if v is not None else "",
                "note": str(n).strip() if n is not None else "",
            }
            col += 2
        out[sid] = rec
    wb.close()
    return out


def _backfill_from_reference(
    merged: OrderedDict[str, dict[str, Any]],
    ref: dict[str, dict[str, dict[str, str]]],
) -> tuple[int, int]:
    filled_cells = 0
    touched_samples = 0
    for sid, rec in merged.items():
        rs = ref.get(sid)
        if not rs:
            continue
        sample_touched = False
        for m in OUTPUT_MARKERS:
            cur = rec.get(m) or {"value": "", "note": ""}
            if not isinstance(cur, dict):
                continue
            if _is_empty_marker(cur):
                rv = (rs.get(m, {}).get("value") or "").strip()
                rn = (rs.get(m, {}).get("note") or "").strip()
                if rv or rn:
                    cur["value"] = rv
                    cur["note"] = rn
                    rec[m] = cur
                    filled_cells += 1
                    sample_touched = True
        if sample_touched:
            touched_samples += 1
    return touched_samples, filled_cells


def parse_ocr_line(line: str) -> dict[str, Any] | None:  # noqa: PLR0915 - TODO: 下个迭代重构 # noqa: PLR0912 - TODO: 下个迭代重构 # noqa: PLR0911 - TODO: 下个迭代重构
    """
    OCR 行解析（尽量鲁棒）。

    1) 若 OCR 识别出 `|` 分隔，则优先用原逻辑（五项各一对，共 12 段以上）。
    2) 若没有 `|` 或段数不足：使用正则提取「样品ID」与多组「value + 阴/阳」对，
       按出现顺序填充 `OUTPUT_MARKERS`（缺项留空）。
    """
    line = (line or "").strip()
    if not line:
        return None

    # 1) 采用 `|` 分隔的结构化解析（尽可能保真）
    if "|" in line:
        parts = [p.strip() for p in line.split("|")]
        if len(parts) >= 12:
            try:
                int(parts[0])
            except ValueError:
                return None
            sid = _normalize_sample_id(parts[1])
            if not sid or not re.search(r"-", sid):
                return None
            rest = parts[2:]
            if len(rest) < 10:
                # 分隔但字段不完整，交给后面的鲁棒解析
                sid_m = re.search(r"([\w\d]+-[\w\d]+-[A-Za-z0-9]+)", line)
                if not sid_m:
                    return None
            else:
                rec: dict[str, Any] = {"样品ID": sid}
                for m in OUTPUT_MARKERS:
                    rec[m] = {"value": "", "note": ""}
                for i, mk in enumerate(OUTPUT_MARKERS):
                    rec[mk] = {
                        "value": _clean_ocr_value(rest[i * 2]),
                        "note": _clean_ocr_note(rest[i * 2 + 1]),
                    }
                return rec

    # 2) 鲁棒解析：先找样品ID（允许 OCR 把 '-' 识别成空格/长横线/下划线）
    # 例如：`725-M1-a`、`725 - M1 - a`、`725—M1_a` 等
    sid_m = re.search(r"([\w\d]+)[\s\-—–_]+([\w\d]+)[\s\-—–_]+([A-Za-z0-9]+)", line)
    if not sid_m:
        return None
    sid = f"{sid_m.group(1)}-{sid_m.group(2)}-{sid_m.group(3)}"
    sid = re.sub(r"[^A-Za-z0-9\-\_]+", "", sid).replace("_", "-")
    sid = re.sub(r"-{2,}", "-", sid).strip("-")
    sid = _normalize_sample_id(sid)
    if not sid or "-" not in sid:
        return None

    # 再找 value+note 成对出现：例如 `2.95 阴性`、`0.00 阳性`、`<2.00 阴性`
    # 注意：OCR 可能在 value 与 note 之间插入 `mIU/ml`、`IU/ml`、`S/CO` 等杂项，
    # 因此改用“按出现顺序的 token 状态机”而不是强依赖相邻。
    val_pat = r"(?:[<≤]?\s*\d+(?:\.\d+)?|[≥>]\s*\d+(?:\.\d+)?)"
    note_pat = r"(阴性|陰性|阳性|陽性|阴|陽|阳|反应|可疑|待查)"
    token_re = re.compile(rf"(?P<val>{val_pat})|(?P<note>{note_pat})")

    def _norm_note(n: str) -> str:
        n = _norm_cell(n)
        if n in ("阴", "陰性"):
            return "阴性"
        if n in ("阳", "陽", "陽性"):
            return "阳性"
        return n

    # 去掉样品ID文本，避免把 sid 的数字误当作 value
    wo_sid = line
    if sid_m:
        a, b = sid_m.span()
        wo_sid = line[:a] + " " + line[b:]

    pairs: list[tuple[str, str]] = []
    cur_val: str = ""
    for m in token_re.finditer(wo_sid):
        v = m.groupdict().get("val")
        nt_raw = m.groupdict().get("note")
        if v:
            cur_val = _clean_ocr_value(v)
        elif nt_raw:
            nt = _norm_note(nt_raw)
            if cur_val and nt:
                pairs.append((cur_val, nt))
                cur_val = ""

    if not pairs:
        return None

    rec: dict[str, Any] = {"样品ID": sid}
    for m in OUTPUT_MARKERS:
        rec[m] = {"value": "", "note": ""}

    for i, (v, nt) in enumerate(pairs[: len(OUTPUT_MARKERS)]):
        rec[OUTPUT_MARKERS[i]] = {"value": v, "note": nt}
    return rec


def extract_rows_ocr_pdf(
    pdf_path: Path,
    *,
    dpi: int = 120,
    lang: str = "chi_sim+eng",
) -> list[dict[str, Any]]:
    if not _OCR_AVAILABLE:
        logger.warning("未安装 pytesseract 或 Pillow，跳过 OCR: %s", pdf_path.name)
        return []
    if fitz is None:
        logger.warning("未安装 PyMuPDF，跳过 OCR: %s", pdf_path.name)
        return []

    out: list[dict[str, Any]] = []
    doc = fitz.open(pdf_path)
    n_pages = doc.page_count
    try:
        for pno in range(n_pages):
            page = doc[pno]
            if pno and pno % 25 == 0:
                logger.info("OCR 进度 %s: %s/%s 页", pdf_path.name, pno, n_pages)
            layer = page.get_text() or ""
            # 已有可检索五项文本的页视为矢量内容，不重复 OCR；仅有封面说明等文字则跳过
            if layer.strip():
                if "Anti-HBs" in layer or "HBsAg" in layer:
                    continue
                continue
            pix = page.get_pixmap(dpi=dpi, alpha=False)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            try:
                txt = pytesseract.image_to_string(img, lang=lang)
            except Exception as e:
                logger.debug("OCR 失败 p%s %s: %s", pno + 1, pdf_path.name, e)
                continue
            if "Anti-HBs" not in txt and "HBsAg" not in txt:
                continue
            for raw_line in txt.splitlines():
                line = raw_line.strip()
                if not line:
                    continue
                rec = parse_ocr_line(line)
                if rec:
                    out.append(rec)
    finally:
        doc.close()
    return out


def merge_records(
    pdf_to_rows: list[tuple[Path, list[dict[str, Any]]]],
    *,
    prefer_last: bool = True,
) -> OrderedDict[str, dict[str, Any]]:
    merged: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for path, rows in pdf_to_rows:
        for rec in rows:
            sid = canonical_sample_id(rec.get("样品ID") or "")
            if not sid:
                continue
            if sid not in merged:
                merged[sid] = {"样品ID": sid, **{m: {"value": "", "note": ""} for m in OUTPUT_MARKERS}}
            cur = merged[sid]
            for m in OUTPUT_MARKERS:
                src = rec.get(m)
                if not isinstance(src, dict):
                    continue
                nv = (src.get("value") or "").strip()
                nn = (src.get("note") or "").strip()
                if nv and (prefer_last or not (str(cur[m].get("value", "")).strip())):
                    cur[m]["value"] = nv
                if nn and (prefer_last or not (str(cur[m].get("note", "")).strip())):
                    cur[m]["note"] = nn
        logger.info("文件 %s 解析 %s 行", path.name, len(rows))
    return merged


def write_excel(merged: OrderedDict[str, dict[str, Any]], out_path: Path) -> None:  # noqa: PLR0915 - TODO: 下个迭代重构
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "血清检测汇总"
    thin = Side(style="thin", color="000000")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A1:A2")
    c = ws["A1"]
    c.value = "样品 ID"
    c.font = Font(bold=True)
    c.alignment = center
    c.border = grid

    col_idx = 2
    for m in OUTPUT_MARKERS:
        c0 = get_column_letter(col_idx)
        c1 = get_column_letter(col_idx + 1)
        ws.merge_cells(f"{c0}1:{c1}1")
        top = ws[f"{c0}1"]
        top.value = m
        top.font = Font(bold=True)
        top.alignment = center
        top.border = grid
        u = ws[f"{c0}2"]
        u.value = MARKER_UNITS[m]
        u.alignment = center
        u.border = grid
        n = ws[f"{c1}2"]
        n.value = "说明"
        n.alignment = center
        n.border = grid
        col_idx += 2

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20

    r = 3
    for sid, rec in merged.items():
        ws.cell(row=r, column=1, value=sid).alignment = center
        ws.cell(row=r, column=1).border = grid
        cc = 2
        for m in OUTPUT_MARKERS:
            blk = rec.get(m) or {}
            v = blk.get("value", "") if isinstance(blk, dict) else ""
            nn = blk.get("note", "") if isinstance(blk, dict) else ""
            cva = ws.cell(row=r, column=cc, value=v)
            cva.alignment = center
            cva.border = grid
            cnb = ws.cell(row=r, column=cc + 1, value=nn)
            cnb.alignment = center
            cnb.border = grid
            cc += 2
        r += 1

    # 列宽
    ws.column_dimensions["A"].width = 14
    for col in range(2, 2 + len(OUTPUT_MARKERS) * 2):
        ws.column_dimensions[get_column_letter(col)].width = 12

    wb.save(out_path)


def discover_pdfs(input_dir: Path) -> list[Path]:
    return sorted(input_dir.glob("*.pdf"), key=lambda p: p.name.lower())


def run(
    input_dir: Path,
    output_path: Path,
    *,
    use_ocr: bool,
    ocr_dpi: int,
    prefer_last: bool,
    reference_excel: Path | None,
    overwrite: bool,
) -> int:
    pdfs = discover_pdfs(input_dir)
    if not pdfs:
        logger.error("目录中无 PDF: %s", input_dir)
        return 1

    pdf_rows: list[tuple[Path, list[dict[str, Any]]]] = []
    for pdf in pdfs:
        digital = extract_rows_digital_pdf(pdf)
        if digital:
            rows = digital
        elif use_ocr:
            rows = extract_rows_ocr_pdf(pdf, dpi=ocr_dpi)
        else:
            rows = []
            try:
                no_tables = not extract_tables_from_pdf(pdf)
            except Exception:
                logger.debug("二次判定表格存在性失败: file=%s", pdf.name, exc_info=True)
                no_tables = True
            if no_tables:
                logger.warning(
                    "未从 %s 解析到表格；扫描件请加 --ocr（需本机 Tesseract + chi_sim+eng）",
                    pdf.name,
                )
        pdf_rows.append((pdf, rows))

    if output_path.exists() and not overwrite:
        logger.error("输出已存在，请使用 --overwrite: %s", output_path)
        return 1

    merged = merge_records(pdf_rows, prefer_last=prefer_last)
    if reference_excel:
        if not reference_excel.exists():
            logger.warning("参考 Excel 不存在，跳过回填：%s", reference_excel)
        else:
            ref = _load_reference_excel(reference_excel)
            touched_samples, filled_cells = _backfill_from_reference(merged, ref)
            logger.info(
                "参考回填完成：样品 %s 个，指标块 %s 个（来源：%s）",
                touched_samples,
                filled_cells,
                reference_excel,
            )
    write_excel(merged, output_path)
    logger.info("已写入 %s 行 -> %s", len(merged), output_path)
    logger.info("完成: samples=%s output=%s", len(merged), output_path)
    return 0


def main() -> None:
    parser = argparse.ArgumentParser(description="中检院血清样本检测报告 PDF 转 Excel")
    parser.add_argument(
        "--input",
        "-i",
        type=Path,
        default=MODULE_DIR / "input",
        help="PDF 输入目录（默认模块下 input）",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        default=MODULE_DIR / "output" / "serology_report_merged.xlsx",
        help="输出 Excel 路径",
    )
    parser.add_argument("--overwrite", action="store_true", help="覆盖已存在输出文件")
    parser.add_argument(
        "--ocr",
        action="store_true",
        help="对无矢量表的扫描 PDF 使用 Tesseract OCR（较慢）",
    )
    parser.add_argument("--ocr-dpi", type=int, default=120, help="OCR 渲染 DPI（默认 120）")
    parser.add_argument(
        "--prefer-first",
        action="store_true",
        help="同一样品 ID 多文件冲突时保留较早文件（默认保留较晚文件）",
    )
    parser.add_argument(
        "--reference-excel",
        type=Path,
        default=None,
        help="可选：参考 Excel（如 09_Word_All_Tables_to_Excel/output/word_tables_merged.xlsx），用于回填 PDF 侧空缺指标",  # noqa: E501
    )
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s %(message)s",
    )

    code = run(
        args.input.resolve(),
        args.output.resolve(),
        use_ocr=args.ocr,
        ocr_dpi=args.ocr_dpi,
        prefer_last=not args.prefer_first,
        reference_excel=args.reference_excel.resolve() if args.reference_excel else None,
        overwrite=args.overwrite,
    )
    raise SystemExit(code)


if __name__ == "__main__":
    main()
