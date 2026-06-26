# -*- coding: utf-8 -*-
"""
将 input/ 下多个 Word(.doc/.docx/.rtf) 的相关检验表合并为一个 Excel 列表。

输出结构参考常见血清学列表：
- 样品ID
- Anti-HBs / HBsAg / Anti-HBc / Anti-HBe / HBeAg（每项两列：数值、说明）

说明：
- 若某些文档仅包含前两项（Anti-HBs/HBsAg），其余指标自动留空。
- 解析依赖 Word COM（pywin32），兼容 Microsoft Office。
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import OrderedDict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

try:
    from docx import Document  # type: ignore
except Exception:
    Document = None  # type: ignore

MODULE_DIR = Path(__file__).resolve().parent
ROOT_DIR = MODULE_DIR.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src.serology_utils import OUTPUT_MARKERS as MARKERS, MARKER_UNITS, canonical_sample_id  # noqa: E402 — 须在 sys.path.insert 之后

WORD_SUFFIXES = {".doc", ".docx", ".rtf"}


def _clean_text(x: Any) -> str:
    if x is None:
        return ""
    s = str(x)
    # Word 单元格结束符
    s = s.replace("\x07", "").replace("\r", "").replace("\x0b", "")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _norm_key(s: str) -> str:
    s = _clean_text(s).lower()
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", s)


def _marker_from_cell(cell: str) -> str | None:
    s = _norm_key(cell)
    if not s:
        return None
    if "antihbs" in s:
        return "Anti-HBs"
    if "hbsag" in s:
        return "HBsAg"
    if "antihbc" in s:
        return "Anti-HBc"
    if "antihbe" in s:
        return "Anti-HBe"
    if "hbeag" in s:
        return "HBeAg"
    return None


def _looks_like_sample_id(s: str) -> bool:
    t = _clean_text(s)
    if not t:
        return False
    # 典型样品ID：001-xx-xx / 含字母数字与连字符
    if "-" in t and re.search(r"[A-Za-z0-9]", t):
        return True
    # 退而求其次：非纯数字且长度较短
    if not re.fullmatch(r"\d+(\.\d+)?", t) and 3 <= len(t) <= 40:
        return True
    return False


def _extract_row_values(
    row: list[str], c0: int, c1: int, sub_row: list[str] | None
) -> tuple[str, str]:
    """
    提取某个指标区间里的（数值, 说明）。
    若子表头存在“说明”列，优先按该列切分。
    """
    seg = [_clean_text(row[i]) if i < len(row) else "" for i in range(c0, c1)]
    if not seg:
        return "", ""

    if sub_row:
        sub = [
            _clean_text(sub_row[i]) if i < len(sub_row) else "" for i in range(c0, c1)
        ]
        note_idx = None
        for i, sh in enumerate(sub):
            if sh == "说明":
                note_idx = i
                break
        if note_idx is not None:
            note = _clean_text(seg[note_idx]) if note_idx < len(seg) else ""
            vals = [x for i, x in enumerate(seg[:note_idx]) if x]
            value = vals[0] if vals else ""
            return value, note

    # 无说明列：若最后一个像“阴性/阳性”则当说明
    non_empty = [x for x in seg if x]
    if not non_empty:
        return "", ""
    if len(non_empty) >= 2 and non_empty[-1] in {
        "阴性",
        "阳性",
        "可疑",
        "反应",
        "待查",
    }:
        return non_empty[0], non_empty[-1]
    return non_empty[0], ""


def _parse_table_matrix(matrix: list[list[str]]) -> list[dict[str, Any]]:
    if not matrix or len(matrix) < 2:
        return []

    # 在前几行找表头：既包含样品ID提示，又包含至少一个 marker
    header_idx = None
    for r in range(min(6, len(matrix))):
        row = matrix[r]
        row_text = " ".join(_clean_text(c) for c in row[:4])
        has_sid = ("样品" in row_text and "ID" in row_text) or (
            "sample" in row_text.lower() and "id" in row_text.lower()
        )
        has_marker = any(_marker_from_cell(c) for c in row)
        if has_sid and has_marker:
            header_idx = r
            break
    if header_idx is None:
        return []

    header = matrix[header_idx]
    sub_row = matrix[header_idx + 1] if header_idx + 1 < len(matrix) else None

    starts: list[tuple[int, str]] = []
    for i, c in enumerate(header):
        mk = _marker_from_cell(c)
        if mk:
            # 同一指标常在“数值列/说明列”重复出现，保留这一段的起始列即可
            if not starts or starts[-1][1] != mk:
                starts.append((i, mk))
    if not starts:
        return []

    starts.sort(key=lambda x: x[0])
    width = max(len(r) for r in matrix)
    groups: list[tuple[str, int, int]] = []
    for i, (st, mk) in enumerate(starts):
        ed = starts[i + 1][0] if i + 1 < len(starts) else width
        groups.append((mk, st, ed))

    data_start = header_idx + 1
    if sub_row and any(
        _clean_text(c) in {"mIU/ml", "IU/ml", "S/CO", "说明"} for c in sub_row
    ):
        data_start = header_idx + 2

    out: list[dict[str, Any]] = []
    for r in range(data_start, len(matrix)):
        row = matrix[r]
        if not any(_clean_text(c) for c in row):
            continue
        sid = ""
        if len(row) > 1 and _looks_like_sample_id(row[1]):
            sid = _clean_text(row[1])
        elif len(row) > 0 and _looks_like_sample_id(row[0]):
            sid = _clean_text(row[0])
        if not sid:
            continue

        rec: dict[str, Any] = {"样品ID": sid}
        for mk in MARKERS:
            rec[mk] = {"value": "", "note": ""}
        for mk, c0, c1 in groups:
            v, n = _extract_row_values(row, c0, c1, sub_row)
            rec[mk] = {"value": v, "note": n}
        out.append(rec)
    return out


def _extract_matrix_from_word_table(t: Any) -> list[list[str]]:
    nrows = int(t.Rows.Count)
    ncols = int(t.Columns.Count)
    matrix: list[list[str]] = []
    for r in range(1, nrows + 1):
        row_vals: list[str] = []
        for c in range(1, ncols + 1):
            try:
                raw = t.Cell(r, c).Range.Text
            except Exception:
                raw = ""
            row_vals.append(_clean_text(raw))
        matrix.append(row_vals)
    return matrix


def _extract_docx_tables(path: Path) -> list[list[list[str]]]:
    if Document is None:
        raise RuntimeError("python-docx 未安装，无法使用 docx 直读")
    doc = Document(str(path))
    out: list[list[list[str]]] = []
    for tb in doc.tables:
        matrix: list[list[str]] = []
        for row in tb.rows:
            matrix.append([_clean_text(c.text) for c in row.cells])
        out.append(matrix)
    return out


class WordComRunner:
    def __init__(self) -> None:
        import win32com.client  # type: ignore

        self.win32 = win32com.client
        self.app = None
        self.doc = None

    def __enter__(self) -> "WordComRunner":
        self.app = self.win32.Dispatch("Word.Application")
        self.app.Visible = False
        try:
            self.app.DisplayAlerts = 0
        except Exception:
            pass
        try:
            self.app.AutomationSecurity = 3
        except Exception:
            pass
        return self

    def open_doc(self, path: Path) -> Any:
        if self.doc is not None:
            try:
                self.doc.Close(False)
            except Exception:
                pass
        try:
            self.doc = self.app.Documents.Open(str(path), False, True, False)
        except Exception:
            self.doc = self.app.Documents.Open(str(path))
        return self.doc

    def __exit__(self, exc_type, exc, tb) -> None:
        if self.doc is not None:
            try:
                self.doc.Close(False)
            except Exception:
                pass
            self.doc = None
        if self.app is not None:
            try:
                self.app.Quit()
            except Exception:
                pass
            self.app = None


def _iter_word_files(input_dir: Path) -> list[Path]:
    files = [
        p
        for p in input_dir.iterdir()
        if p.is_file()
        and p.suffix.lower() in WORD_SUFFIXES
        and not p.name.startswith("~$")
    ]
    files.sort(key=lambda x: x.name.lower())
    return files


def _merge_records(records: list[dict[str, Any]]) -> OrderedDict[str, dict[str, Any]]:
    merged: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for rec in records:
        sid = canonical_sample_id(rec.get("样品ID"))
        if not sid:
            continue
        if sid not in merged:
            merged[sid] = {
                "样品ID": sid,
                **{m: {"value": "", "note": ""} for m in MARKERS},
            }
        cur = merged[sid]
        for mk in MARKERS:
            blk = rec.get(mk) if isinstance(rec.get(mk), dict) else {}
            nv = _clean_text(blk.get("value"))
            nn = _clean_text(blk.get("note"))
            # 仅在当前为空时填充，避免后续错误覆盖
            if nv and not _clean_text(cur[mk].get("value")):
                cur[mk]["value"] = nv
            if nn and not _clean_text(cur[mk].get("note")):
                cur[mk]["note"] = nn
    return merged


def _load_reference_excel(path: Path) -> dict[str, dict[str, dict[str, str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out: dict[str, dict[str, dict[str, str]]] = {}
    for r in range(3, ws.max_row + 1):
        sid_raw = ws.cell(r, 1).value
        sid = canonical_sample_id(str(sid_raw) if sid_raw is not None else "")
        if not sid:
            continue
        rec: dict[str, dict[str, str]] = {}
        col = 2
        for mk in MARKERS:
            v = ws.cell(r, col).value
            n = ws.cell(r, col + 1).value
            rec[mk] = {
                "value": _clean_text(v),
                "note": _clean_text(n),
            }
            col += 2
        out[sid] = rec
    wb.close()
    return out


def _backfill_from_reference(
    merged: OrderedDict[str, dict[str, Any]],
    ref: dict[str, dict[str, dict[str, str]]],
) -> tuple[int, int]:
    filled_blocks = 0
    touched_samples = 0
    for sid, rec in merged.items():
        rs = ref.get(sid)
        if not rs:
            continue
        touched = False
        for mk in MARKERS:
            cur = rec.get(mk) if isinstance(rec.get(mk), dict) else {}
            if not isinstance(cur, dict):
                continue
            cv = _clean_text(cur.get("value"))
            cn = _clean_text(cur.get("note"))
            if cv or cn:
                continue
            rv = _clean_text(rs.get(mk, {}).get("value"))
            rn = _clean_text(rs.get(mk, {}).get("note"))
            if rv or rn:
                cur["value"] = rv
                cur["note"] = rn
                rec[mk] = cur
                filled_blocks += 1
                touched = True
        if touched:
            touched_samples += 1
    return touched_samples, filled_blocks


def _write_output(merged: OrderedDict[str, dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A1:A2")
    ws["A1"] = "样品 ID"
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = center
    ws["A1"].border = border

    col = 2
    for mk in MARKERS:
        c0 = get_column_letter(col)
        c1 = get_column_letter(col + 1)
        ws.merge_cells(f"{c0}1:{c1}1")
        ws[f"{c0}1"] = mk
        ws[f"{c0}1"].font = Font(bold=True)
        ws[f"{c0}1"].alignment = center
        ws[f"{c0}1"].border = border

        ws[f"{c0}2"] = MARKER_UNITS[mk]
        ws[f"{c0}2"].alignment = center
        ws[f"{c0}2"].border = border

        ws[f"{c1}2"] = "说明"
        ws[f"{c1}2"].alignment = center
        ws[f"{c1}2"].border = border
        col += 2

    r = 3
    for sid, rec in merged.items():
        ws.cell(r, 1, sid).alignment = center
        ws.cell(r, 1).border = border
        c = 2
        for mk in MARKERS:
            v = _clean_text(rec[mk].get("value"))
            n = _clean_text(rec[mk].get("note"))
            ws.cell(r, c, v).alignment = center
            ws.cell(r, c).border = border
            ws.cell(r, c + 1, n).alignment = center
            ws.cell(r, c + 1).border = border
            c += 2
        r += 1

    ws.column_dimensions["A"].width = 14
    for c in range(2, 12):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.freeze_panes = "A3"
    wb.save(output_path)
    wb.close()


def run(
    input_dir: Path,
    output_path: Path,
    *,
    prefer_docx_reader: bool = True,
    reference_excel: Path | None = None,
) -> tuple[int, int, int]:
    files = _iter_word_files(input_dir)
    if not files:
        raise FileNotFoundError(f"输入目录未找到 Word/RTF 文件：{input_dir}")

    all_records: list[dict[str, Any]] = []
    parsed_tables = 0
    com_candidates: list[Path] = []
    for fp in files:
        if prefer_docx_reader and fp.suffix.lower() == ".docx" and Document is not None:
            try:
                for matrix in _extract_docx_tables(fp):
                    rows = _parse_table_matrix(matrix)
                    if rows:
                        parsed_tables += 1
                        all_records.extend(rows)
                continue
            except Exception:
                # docx 直读失败时回退 COM
                com_candidates.append(fp)
                continue
        com_candidates.append(fp)

    if com_candidates:
        with WordComRunner() as runner:
            for fp in com_candidates:
                doc = runner.open_doc(fp)
                tables = doc.Content.Tables
                for i in range(1, int(tables.Count) + 1):
                    t = tables.Item(i)
                    matrix = _extract_matrix_from_word_table(t)
                    rows = _parse_table_matrix(matrix)
                    if rows:
                        parsed_tables += 1
                        all_records.extend(rows)

    merged = _merge_records(all_records)
    if reference_excel:
        if reference_excel.exists():
            ref = _load_reference_excel(reference_excel)
            _backfill_from_reference(merged, ref)
        else:
            print(f"[警告] 参考 Excel 不存在，跳过回填：{reference_excel}")
    _write_output(merged, output_path)
    return len(files), parsed_tables, len(merged)


def main() -> None:
    parser = argparse.ArgumentParser(description="将多个 Word 表格合并为一个列表 Excel")
    parser.add_argument(
        "--input-dir", "-i", default=str(MODULE_DIR / "input"), help="输入目录"
    )
    parser.add_argument(
        "--output",
        "-o",
        default=str(MODULE_DIR / "output" / "word_tables_merged.xlsx"),
        help="输出 xlsx 路径",
    )
    parser.add_argument(
        "--disable-docx-reader",
        action="store_true",
        help="禁用 docx 直读，强制全部走 Word COM（一般不推荐）",
    )
    parser.add_argument(
        "--reference-excel",
        default=None,
        help="可选：参考 Excel（如 12_PDF_Batch_to_Excel/output/serology_report_merged.xlsx），用于回填 Word 侧空缺指标",
    )
    args = parser.parse_args()

    input_dir = Path(args.input_dir).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    ref_excel = (
        Path(args.reference_excel).expanduser().resolve()
        if args.reference_excel
        else None
    )
    n_files, n_tables, n_rows = run(
        input_dir,
        output_path,
        prefer_docx_reader=not bool(args.disable_docx_reader),
        reference_excel=ref_excel,
    )
    print(
        f"完成：文件 {n_files} 个，命中表 {n_tables} 张，汇总样品 {n_rows} 行 -> {output_path}"
    )


if __name__ == "__main__":
    main()
