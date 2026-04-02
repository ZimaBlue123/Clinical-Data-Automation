# -*- coding: utf-8 -*-
"""
Word 表格导出为 Excel（精准表头对齐 + 兼容性优先）。

设计点：
- 用 Word COM 读取 doc.Content.Tables（避免 python-docx 对复杂表格/RTF 的限制）
- 支持按 table index 或按表头关键字筛选目标表
- 支持多行表头合并（header_rows），写入 Excel 时冻结窗格、加粗表头、自动换行
- 大表优先按「行」取 Rows(r).Range.Text 再切分单元格，避免整表 Range.Text 截断与逐格 COM 过慢
- Document.Tables 与 document.xml 中 <w:tbl> 数量可能不一致（嵌套表）；选表请用 --list-word-tables

架构分层（自上而下）：
1) CLI（argparse）→ 解析参数、list/dry-run 分支
2) export_word_tables_to_excel：编排 Word 会话、选表、写 xlsx
3) 选表：_collect_selected_tables（合并 / 表题 / 序号 / 全量+关键字）
4) 抽取：_extract_table（整表 Range.Text 小表快路径 → 按行 Range.Text → 按行 Cells → Cell(r,c)）
5) 写出：_write_workbook_for_tables（openpyxl + 样式）
"""

from __future__ import annotations

import argparse
import re
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


WORD_SUFFIXES = {".doc", ".docx", ".rtf"}

# 整表 Range.Text 易截断；超过此行数或文本长度则走按行抽取
FAST_PATH_MAX_TABLE_ROWS = 300
WHOLE_TABLE_RANGE_TEXT_MAX_CHARS = 32000
# 快速路径：解析出的行列与 Word 报告差异超过阈值则放弃快路径（在 _extract_table 内）
FAST_PATH_ROW_RATIO = 0.8
FAST_PATH_COL_RATIO = 0.8


def _strip_word_cell_text(text: Any) -> str:
    if text is None:
        return ""
    s = str(text)
    # Word 单元格常见结束符
    s = s.replace("\x07", "")  # cell mark
    s = s.replace("\r", "")  # paragraph mark
    s = s.replace("\x0b", "")
    s = s.strip()
    # 折叠多空白
    s = re.sub(r"\s+", " ", s)
    return s


def _norm_key(s: str) -> str:
    # 用于关键字匹配：去空白/符号、统一大小写
    s = _strip_word_cell_text(s)
    s = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "", s)
    return s.upper()


@dataclass(frozen=True)
class TableData:
    table_index: int
    rows: list[list[str]]  # 纯文本矩阵（已清洗）
    merged_range: Optional[tuple[int, int]] = None  # (from_index, to_index) 合并多段 Word 表时记录

    @property
    def nrows(self) -> int:
        return len(self.rows)

    @property
    def ncols(self) -> int:
        return max((len(r) for r in self.rows), default=0)


def _cells_from_word_line(ln: str) -> list[str]:
    """将 Word 表格内一行原始文本（单元格以 \\x07 分隔）解析为清洗后的单元格列表。"""
    parts = [p for p in ln.split("\x07")]
    if parts and parts[-1] == "":
        parts = parts[:-1]
    return [_strip_word_cell_text(p) for p in parts]


def _matrix_from_whole_table_text(raw: str) -> list[list[str]]:
    """整表 Range.Text → 行矩阵（不含列对齐）。"""
    raw = raw.replace("\x0b", "")
    matrix: list[list[str]] = []
    for ln in raw.split("\r"):
        if not ln:
            continue
        matrix.append(_cells_from_word_line(ln))
    return matrix


def _pad_rows_to_ncols(rows: list[list[str]], ncols: int) -> list[list[str]]:
    out: list[list[str]] = []
    for r in rows:
        if len(r) >= ncols:
            out.append(r[:ncols])
        else:
            out.append(r + [""] * (ncols - len(r)))
    return out


def _docx_approx_table_row_counts(path: Path) -> list[int]:
    """
    仅从 docx(zip) 的 document.xml 统计每个 <w:tbl> 内 <w:tr> 数量（不启动 Word）。
    注意：XML 中嵌套表格会单独计数，数量常大于 Word 的 Document.Tables.Count；
    选表序号请以 --list-word-tables（COM）为准。
    """
    if path.suffix.lower() != ".docx":
        return []
    try:
        with zipfile.ZipFile(path) as z:
            xml = z.read("word/document.xml").decode("utf-8", errors="replace")
    except Exception:
        return []
    parts = xml.split("<w:tbl")
    counts: list[int] = []
    for part in parts[1:]:
        seg = part.split("</w:tbl>", 1)[0]
        counts.append(seg.count("<w:tr"))
    return counts


def _self_check_table(t: Any, table_index: int, matrix: list[list[str]], source: str) -> None:
    try:
        er = int(t.Rows.Count)
        ec = int(t.Columns.Count)
    except Exception:
        return
    mr = len(matrix)
    mc = max((len(r) for r in matrix), default=0)
    ok_r = mr == er
    ok_c = mc == ec or (ec > 0 and abs(mc - ec) <= max(2, ec // 20))
    if not ok_r or not ok_c:
        print(
            f"[自检] 表 {table_index} ({source}): Word 报告 {er} 行 x {ec} 列；"
            f"导出 {mr} 行 x {mc} 列 —— 可能不完整或含复杂合并单元格。",
            file=sys.stderr,
        )


def _extract_table_by_row_range_text(t: Any) -> list[list[str]]:
    """
    按行取 Rows(r).Range.Text 再按 \\x07 切分单元格。
    大表时 COM 调用次数约为「行数」，远少于逐单元格（行×列），且不易触发整表 Range.Text 截断。
    """
    nrows = int(t.Rows.Count)
    ncols = int(t.Columns.Count)
    matrix: list[list[str]] = []
    for r in range(1, nrows + 1):
        raw = str(t.Rows(r).Range.Text or "").replace("\x0b", "")
        row_vals: list[str] = []
        for ln in raw.split("\r"):
            if not ln:
                continue
            row_vals.extend(_cells_from_word_line(ln))
        if len(row_vals) < ncols:
            row_vals = row_vals + [""] * (ncols - len(row_vals))
        elif len(row_vals) > ncols:
            row_vals = row_vals[:ncols]
        matrix.append(row_vals)
    return matrix


def _extract_table_by_row_cells(t: Any) -> list[list[str]]:
    """按行遍历 Cells（对合并单元格通常比 Cell(r,c) 双循环更稳）。"""
    nrows = int(t.Rows.Count)
    ncols = int(t.Columns.Count)
    matrix: list[list[str]] = []
    for r in range(1, nrows + 1):
        row_obj = t.Rows(r)
        ncells = int(row_obj.Cells.Count)
        row_vals: list[str] = []
        for c in range(1, ncells + 1):
            try:
                raw2 = row_obj.Cells(c).Range.Text
            except Exception:
                raw2 = None
            row_vals.append(_strip_word_cell_text(raw2))
        # 与 Columns.Count 对齐：合并格可能导致列数与逻辑列不完全一致，取较大者再裁切/填充
        if len(row_vals) < ncols:
            row_vals = row_vals + [""] * (ncols - len(row_vals))
        elif len(row_vals) > ncols:
            row_vals = row_vals[:ncols]
        matrix.append(row_vals)
    return matrix


def _extract_table(doc: Any, table_index: int, *, quiet: bool = False) -> TableData:
    t = doc.Content.Tables.Item(table_index)
    expected_rows = int(t.Rows.Count)
    expected_cols = int(t.Columns.Count)

    # 大表或超长文本：Word 整表 Range.Text 可能被截断，直接跳过快速路径
    try:
        raw_probe = str(t.Range.Text or "")
    except Exception:
        raw_probe = ""
    skip_fast = expected_rows > FAST_PATH_MAX_TABLE_ROWS or len(raw_probe) >= WHOLE_TABLE_RANGE_TEXT_MAX_CHARS

    # 快速路径：一次性取 Table.Range.Text 再切分（小表更快）
    if not skip_fast:
        try:
            matrix = _matrix_from_whole_table_text(raw_probe)
            if matrix:
                parsed_rows = len(matrix)
                parsed_cols = max((len(r) for r in matrix), default=0)
                rows_ok = parsed_rows >= max(1, int(expected_rows * FAST_PATH_ROW_RATIO))
                cols_ok = parsed_cols >= max(1, int(expected_cols * FAST_PATH_COL_RATIO))
                if rows_ok and cols_ok:
                    if not quiet:
                        _self_check_table(t, table_index, matrix, "Range.Text 快速路径")
                    return TableData(table_index=table_index, rows=_pad_rows_to_ncols(matrix, expected_cols))
        except Exception:
            pass

    # 主路径：按行 Range.Text（大表首选，COM 次数≈行数）
    matrix2: list[list[str]] = []
    try:
        matrix2 = _extract_table_by_row_range_text(t)
    except Exception:
        matrix2 = []

    # 行数一致即可采用（列宽用 _pad_rows_to_ncols 对齐）；避免因个别合并格列数波动退回逐格 COM
    ok_shape = len(matrix2) == expected_rows and bool(matrix2)
    if not ok_shape:
        try:
            matrix2 = _extract_table_by_row_cells(t)
        except Exception:
            matrix2 = []

    if len(matrix2) != expected_rows or not matrix2:
        matrix2 = []
        for r in range(1, expected_rows + 1):
            row2: list[str] = []
            for c in range(1, expected_cols + 1):
                try:
                    raw2 = t.Cell(r, c).Range.Text
                except Exception:
                    raw2 = None
                row2.append(_strip_word_cell_text(raw2))
            matrix2.append(row2)

    matrix2 = _pad_rows_to_ncols(matrix2, expected_cols)
    if not quiet:
        _self_check_table(t, table_index, matrix2, "按行Range/逐Cell")
    return TableData(table_index=table_index, rows=matrix2)


def _merge_table_range(doc: Any, start_i: int, end_i: int, *, quiet: bool = False) -> TableData:
    """将多个连续 Word 表纵向拼成一张（跨页被拆成多段 <w:tbl> 时常用）。"""
    all_rows: list[list[str]] = []
    per_counts: list[tuple[int, int]] = []
    for i in range(start_i, end_i + 1):
        td = _extract_table(doc, i, quiet=True)
        per_counts.append((i, td.nrows))
        all_rows.extend(td.rows)
    ncols = max((len(r) for r in all_rows), default=0)
    all_rows = _pad_rows_to_ncols(all_rows, ncols)
    if not quiet:
        total_src = sum(n for _, n in per_counts)
        print(
            f"[合并] Word 表 {start_i}–{end_i} 共 {len(per_counts)} 段，源行数合计 {total_src}；"
            f"合并后 {len(all_rows)} 行 x {ncols} 列。",
            file=sys.stderr,
        )
    return TableData(table_index=start_i, rows=all_rows, merged_range=(start_i, end_i))


def _list_word_tables_via_com(input_path: Path) -> None:
    """用 Word COM 列出顶层表序号及行列数（与 --table-index 一致）。"""
    with WordComRunner() as runner:
        doc = runner.open_doc(input_path)
        ts = doc.Content.Tables
        n = int(ts.Count)
        print("index\trows\tcols")
        for i in range(1, n + 1):
            t = ts.Item(i)
            print(f"{i}\t{int(t.Rows.Count)}\t{int(t.Columns.Count)}")
        print(f"total_word_tables_com\t{n}", file=sys.stderr)


class WordComRunner:
    def __init__(self) -> None:
        try:
            import win32com.client  # type: ignore
        except ImportError as e:
            raise RuntimeError("需要安装 pywin32 才能使用 Word COM") from e
        self.win32 = win32com.client
        self.app = None
        self.doc = None

    def __enter__(self) -> "WordComRunner":
        self.app = self.win32.Dispatch("Word.Application")
        self.app.Visible = False
        try:
            self.app.DisplayAlerts = 0
        except Exception:
            pass
        # 禁用宏（避免安全提示阻塞）
        # msoAutomationSecurityForceDisable = 3
        try:
            self.app.AutomationSecurity = 3
        except Exception:
            pass
        return self

    def open_doc(self, path: Path) -> Any:
        if self.doc is not None:
            try:
                self.doc.Close(False)
            except Exception:
                pass
        # 尽量避免弹窗/阻塞（更新链接、兼容转换等）
        # 说明：不同 Office 版本参数名/位置略有差异；此处使用位置参数以最大化兼容。
        # Documents.Open(FileName, ConfirmConversions, ReadOnly, AddToRecentFiles, PasswordDocument, PasswordTemplate, Revert)
        try:
            self.doc = self.app.Documents.Open(str(path), False, True, False)
        except Exception:
            # 兜底：使用最简 Open
            self.doc = self.app.Documents.Open(str(path))
        return self.doc

    def close(self) -> None:
        if self.doc is not None:
            try:
                self.doc.Close(False)
            except Exception:
                pass
            self.doc = None
        if self.app is not None:
            try:
                self.app.Quit()
            except Exception:
                pass
            self.app = None

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()


def _table_header_text(table: TableData, header_rows: int) -> str:
    # 用前 header_rows 行拼接一个可匹配的 header 文本（用于关键词筛表）
    parts: list[str] = []
    for r in range(min(header_rows, table.nrows)):
        parts.extend([x for x in table.rows[r] if x])
    return " ".join(parts)


def _find_table_index_by_title(doc: Any, title_text: str) -> Optional[int]:
    """
    通过“表题/标题文本”定位表格：
    - 优先返回“包含该文本的 Range 所在的表”
    - 若文本不在表内，尝试返回“该文本之后最近的一个表”
    """
    if not title_text:
        return None

    # 先走“靠近表格的上下文”扫描（更稳定，也避免 Find 在大文档里潜在阻塞）
    try:
        title_norm = _norm_key(title_text)
        # 取更稳的 token：数字段 + 中英文连续串
        toks = re.findall(r"[0-9]+(?:\\.[0-9]+)*|[A-Z]+|[\\u4e00-\\u9fff]+", title_norm)
        toks = [t for t in toks if t]
        if toks:
            all_tables = doc.Content.Tables
            best_i: Optional[int] = None
            best_score = -1
            for i in range(1, int(all_tables.Count) + 1):
                t = all_tables.Item(i)
                ts = int(t.Range.Start)
                # 向前取一段上下文（通常表题在表格前）
                ctx = doc.Range(max(0, ts - 400), ts).Text
                ctx_norm = _norm_key(ctx)
                score = sum(1 for tk in toks if tk in ctx_norm)
                if score > best_score:
                    best_score = score
                    best_i = int(i)
            # 至少命中 2 个 token 才认为可靠（避免仅数字或通用词误命中）
            if best_i is not None and best_score >= min(2, len(toks)):
                return best_i
    except Exception:
        pass

    rng = doc.Content
    find = rng.Find
    find.ClearFormatting()
    find.Text = title_text
    find.Replacement.Text = ""
    find.Forward = True
    find.Wrap = 1  # wdFindContinue
    find.Format = False

    ok = find.Execute()
    if not ok:
        return None

    # 1) 命中点是否在某个 table 内
    try:
        hit_tables = rng.Tables
        if hit_tables is not None and int(hit_tables.Count) > 0:
            t = hit_tables.Item(1)
            # 在 doc.Content.Tables 中找 index
            all_tables = doc.Content.Tables
            for i in range(1, int(all_tables.Count) + 1):
                if all_tables.Item(i).Range.Start == t.Range.Start and all_tables.Item(i).Range.End == t.Range.End:
                    return int(i)
    except Exception:
        pass

    # 2) 不在表内：找命中点之后最近的一个 table
    try:
        start = int(rng.Start)
        all_tables = doc.Content.Tables
        best_i: Optional[int] = None
        best_start: Optional[int] = None
        for i in range(1, int(all_tables.Count) + 1):
            t = all_tables.Item(i)
            ts = int(t.Range.Start)
            if ts >= start and (best_start is None or ts < best_start):
                best_start = ts
                best_i = int(i)
        return best_i
    except Exception:
        return None


def _select_tables(
    all_tables: list[TableData],
    *,
    table_indices: Optional[list[int]],
    header_keywords: Optional[list[str]],
    header_rows: int,
) -> list[TableData]:
    if table_indices:
        wanted = set(table_indices)
        return [t for t in all_tables if t.table_index in wanted]

    if header_keywords:
        keys = [_norm_key(k) for k in header_keywords if _norm_key(k)]
        if not keys:
            return all_tables
        out: list[TableData] = []
        for t in all_tables:
            hdr = _norm_key(_table_header_text(t, header_rows=header_rows))
            if all(k in hdr for k in keys):
                out.append(t)
        return out

    # 默认：全部导出
    return all_tables


def _collect_selected_tables(
    doc: Any,
    *,
    total_tables: int,
    merge_tables_from: Optional[int],
    merge_tables_to: Optional[int],
    table_title: Optional[str],
    table_indices: Optional[list[int]],
    header_keywords: Optional[list[str]],
    header_rows: int,
    quiet: bool,
) -> list[TableData]:
    """在已打开的 doc 上按参数选择并抽取表格（单一职责，供 export 与自检复用）。"""
    if merge_tables_from is not None:
        start_m = int(merge_tables_from)
        end_m = int(merge_tables_to) if merge_tables_to is not None else total_tables
        if start_m < 1 or end_m > total_tables or start_m > end_m:
            raise ValueError(
                f"合并区间无效：{start_m}–{end_m}；文档共有 {total_tables} 个 Word 表（1-based）"
            )
        return [_merge_table_range(doc, start_m, end_m, quiet=quiet)]

    if table_title:
        title_table_index = _find_table_index_by_title(doc, table_title)
        if title_table_index is None:
            raise ValueError(f"未在 Word 中定位到表题：{table_title!r}")
        return [_extract_table(doc, int(title_table_index), quiet=quiet)]

    if table_indices:
        unique_indices = sorted(set(int(i) for i in table_indices))
        bad = [i for i in unique_indices if i < 1 or i > total_tables]
        if bad:
            raise ValueError(f"表格序号越界：{bad}；文档共有 {total_tables} 个表")
        return [_extract_table(doc, i, quiet=quiet) for i in unique_indices]

    all_tables: list[TableData] = []
    for i in range(1, total_tables + 1):
        all_tables.append(_extract_table(doc, i, quiet=quiet))
    return _select_tables(
        all_tables,
        table_indices=table_indices,
        header_keywords=header_keywords,
        header_rows=header_rows,
    )


def _merge_headers(table: TableData, header_rows: int) -> tuple[list[str], list[list[str]]]:
    """
    将多行表头合并为单行列名：例如第1行=分组，第2行=指标 -> "分组 / 指标"
    返回：(merged_headers, body_rows)
    """
    hr = max(1, header_rows)
    hr = min(hr, table.nrows) if table.nrows else 1
    ncols = table.ncols

    # 取 header 矩阵并做“向左/向上填充”以适配 Word 合并单元格带来的空洞
    header = [[(table.rows[r][c] if c < len(table.rows[r]) else "") for c in range(ncols)] for r in range(hr)]

    # 向左填充：同一行空值继承左侧（常见于横向合并）
    for r in range(hr):
        last = ""
        for c in range(ncols):
            if header[r][c]:
                last = header[r][c]
            else:
                header[r][c] = last

    # 向上填充：空值继承上一行（常见于纵向合并）
    for c in range(ncols):
        last = ""
        for r in range(hr):
            if header[r][c]:
                last = header[r][c]
            else:
                header[r][c] = last

    merged: list[str] = []
    for c in range(ncols):
        parts = []
        for r in range(hr):
            v = header[r][c]
            if v and (not parts or parts[-1] != v):
                parts.append(v)
        name = " / ".join(parts).strip()
        merged.append(name if name else f"COL_{c+1}")

    body = table.rows[hr:] if table.nrows >= hr else []
    return merged, body


def _fill_header_matrix_for_style(table: TableData, header_rows: int) -> list[list[str]]:
    """
    生成用于“多行表头美化”的 header 矩阵：做与 _merge_headers 同样的向左/向上填充，
    但保留为多行，写入 Excel 便于冻结窗格与视觉分层。
    """
    hr = max(1, header_rows)
    hr = min(hr, table.nrows) if table.nrows else 1
    ncols = table.ncols
    header = [[(table.rows[r][c] if c < len(table.rows[r]) else "") for c in range(ncols)] for r in range(hr)]

    for r in range(hr):
        last = ""
        for c in range(ncols):
            if header[r][c]:
                last = header[r][c]
            else:
                header[r][c] = last

    for c in range(ncols):
        last = ""
        for r in range(hr):
            if header[r][c]:
                last = header[r][c]
            else:
                header[r][c] = last

    return header


def _safe_sheet_name(name: str, used: set[str]) -> str:
    # Excel sheet name 限制：<=31，不能含 : \ / ? * [ ]
    s = re.sub(r"[:\\\\/?*\\[\\]]+", "_", name).strip()
    s = s[:31] if len(s) > 31 else s
    if not s:
        s = "Table"
    base = s
    i = 2
    while s in used:
        suffix = f"_{i}"
        s = (base[: 31 - len(suffix)] + suffix)[:31]
        i += 1
    used.add(s)
    return s


def _autosize_columns(ws) -> None:
    # 简单按字符宽估算列宽（兼容优先；不依赖 Excel COM）
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)


def _apply_table_style(ws, *, header_rows: int, ncols: int) -> None:
    """
    轻量美化（兼容性优先）：表头深色底+白字、冻结窗格、全表细边框、自动筛选。
    """
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="1F4E79")  # 深蓝
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="top", wrap_text=True)

    max_row = ws.max_row
    max_col = min(ncols, ws.max_column)
    if max_col <= 0 or max_row <= 0:
        return
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.border = border
            if r <= header_rows:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
            else:
                if cell.alignment is None:
                    cell.alignment = body_align
                else:
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical="top",
                        wrap_text=True,
                    )

    # 冻结到表头下方
    ws.freeze_panes = ws.cell(header_rows + 1, 1).coordinate
    # 自动筛选
    header_row_for_filter = max(1, min(header_rows, max_row))
    ws.auto_filter.ref = f"A{header_row_for_filter}:{openpyxl.utils.get_column_letter(max_col)}{max_row}"
    # 表头行高
    for r in range(1, header_rows + 1):
        ws.row_dimensions[r].height = 22


def _write_workbook_for_tables(
    selected: list[TableData],
    output_path: Path,
    *,
    header_rows: int,
    table_title: Optional[str],
) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()

    for t in selected:
        header_matrix = _fill_header_matrix_for_style(t, header_rows=header_rows)
        merged_headers, body = _merge_headers(t, header_rows=header_rows)
        base_name = f"Table_{t.table_index}"
        if t.merged_range:
            a, b = t.merged_range
            base_name = f"Table_{a}_to_{b}"
        if table_title:
            base_name = _strip_word_cell_text(table_title)[:31] or base_name
        sheet_name = _safe_sheet_name(base_name, used_names)
        ws = wb.create_sheet(sheet_name)

        for r in range(len(header_matrix)):
            for c, v in enumerate(header_matrix[r], start=1):
                ws.cell(r + 1, c, value=v)

        start_row = len(header_matrix) + 1
        for r, row in enumerate(body, start=start_row):
            for c in range(1, len(merged_headers) + 1):
                v = row[c - 1] if c - 1 < len(row) else ""
                ws.cell(r, c, value=v)

        _apply_table_style(ws, header_rows=len(header_matrix), ncols=len(merged_headers))
        _autosize_columns(ws)

    wb.save(str(output_path))
    wb.close()


def export_word_tables_to_excel(
    *,
    input_path: Path,
    output_path: Path,
    table_indices: Optional[list[int]],
    header_keywords: Optional[list[str]],
    header_rows: int,
    table_title: Optional[str],
    merge_tables_from: Optional[int] = None,
    merge_tables_to: Optional[int] = None,
    quiet: bool = False,
    dry_run: bool = False,
) -> Path:
    if not input_path.is_file():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")
    if input_path.suffix.lower() not in WORD_SUFFIXES:
        raise ValueError(f"不支持的输入后缀: {input_path.suffix}（仅支持 {sorted(WORD_SUFFIXES)}）")

    with WordComRunner() as runner:
        doc = runner.open_doc(input_path)
        total_tables = int(doc.Content.Tables.Count)
        selected = _collect_selected_tables(
            doc,
            total_tables=total_tables,
            merge_tables_from=merge_tables_from,
            merge_tables_to=merge_tables_to,
            table_title=table_title,
            table_indices=table_indices,
            header_keywords=header_keywords,
            header_rows=header_rows,
            quiet=quiet,
        )

    if not selected:
        raise ValueError("未筛选到任何表格：请检查 --table-indices 或 --header-keywords")

    if dry_run:
        print(f"[dry-run] Word 顶层表数: {total_tables}", file=sys.stderr)
        for t in selected:
            mr = f" merged={t.merged_range}" if t.merged_range else ""
            print(f"[dry-run] 导出块: table_index={t.table_index}{mr} rows={t.nrows} cols={t.ncols}", file=sys.stderr)
        return output_path

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _write_workbook_for_tables(
        selected,
        output_path,
        header_rows=header_rows,
        table_title=table_title,
    )
    return output_path


def _parse_int_list(s: str) -> list[int]:
    out: list[int] = []
    for part in re.split(r"[,\s]+", s.strip()):
        if not part:
            continue
        out.append(int(part))
    return out


def main() -> None:
    parser = argparse.ArgumentParser(description="将 Word 指定表格导出为 Excel（精准表头对齐）")
    default_input = Path(__file__).resolve().parent / "input"
    default_output_dir = Path(__file__).resolve().parent / "output"

    parser.add_argument("--input", "-i", required=True, help="Word/RTF 文件路径（.doc/.docx/.rtf）")
    parser.add_argument("--output", "-o", default=None, help="输出 xlsx 路径（默认 output/<name>_tables.xlsx）")
    parser.add_argument("--table-indices", default=None, help="要导出的表格序号（1-based），例如 1,3,5")
    parser.add_argument("--table-index", type=int, default=None, help="只导出单个表格序号（1-based），例如 9")
    parser.add_argument("--header-keywords", default=None, help="按表头关键字筛选（逗号分隔），例如 系统器官分类,首选术语")
    parser.add_argument("--header-rows", type=int, default=1, help="表头行数（默认 1；用于多级表头合并）")
    parser.add_argument(
        "--table-title",
        default=None,
        help="按“表题/标题文本”定位目标表（优先级最高），例如 表16.2.6.1 基础阶段免疫原性清单(FAS)",
    )
    parser.add_argument(
        "--merge-tables-from",
        type=int,
        default=None,
        help="从第 N 个 Word 表开始纵向合并（含 N）；视觉上跨多页的一张表常被拆成多个 Word 表，需合并导出",
    )
    parser.add_argument(
        "--merge-tables-to",
        type=int,
        default=None,
        help="合并到第 M 个 Word 表（含 M）；默认合并到文档最后一个表",
    )
    parser.add_argument("--quiet", action="store_true", help="减少自检信息输出到 stderr")
    parser.add_argument(
        "--list-docx-tables",
        action="store_true",
        help="仅 .docx：不启动 Word，列出 document.xml 中每段 <w:tbl> 的大致行数（含嵌套表，序号≠ Word 表号）",
    )
    parser.add_argument(
        "--list-word-tables",
        action="store_true",
        help="启动 Word，列出 Document.Tables 的序号及行列数（与 --table-index / 合并参数一致）",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="只抽取并打印行列统计，不写 xlsx（用于大表自检）",
    )

    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()

    if args.list_word_tables and args.list_docx_tables:
        print("请只使用 --list-word-tables 或 --list-docx-tables 之一", file=sys.stderr)
        raise SystemExit(2)

    if args.list_word_tables:
        _list_word_tables_via_com(input_path)
        return

    if args.list_docx_tables:
        counts = _docx_approx_table_row_counts(input_path)
        if not counts:
            print("无法列出表格（需 .docx，或文件无法读取）", file=sys.stderr)
            raise SystemExit(2)
        print("index\tapprox_rows_xml")
        for i, n in enumerate(counts, 1):
            print(f"{i}\t{n}")
        print(f"total_xml_tbl\t{len(counts)}", file=sys.stderr)
        return

    if args.output:
        output_path = Path(args.output).expanduser().resolve()
    else:
        default_output_dir.mkdir(parents=True, exist_ok=True)
        output_path = default_output_dir / f"{input_path.stem}_tables.xlsx"

    table_indices: Optional[list[int]] = _parse_int_list(args.table_indices) if args.table_indices else None
    if args.table_index is not None:
        table_indices = [int(args.table_index)]
    header_keywords: Optional[list[str]] = None
    if args.header_keywords:
        header_keywords = [p.strip() for p in args.header_keywords.split(",") if p.strip()]

    merge_from = args.merge_tables_from
    merge_to = args.merge_tables_to
    table_title = args.table_title
    if merge_from is not None:
        table_indices = None
        table_title = None

    out = export_word_tables_to_excel(
        input_path=input_path,
        output_path=output_path,
        table_indices=table_indices,
        header_keywords=header_keywords,
        header_rows=int(args.header_rows),
        table_title=table_title,
        merge_tables_from=merge_from,
        merge_tables_to=merge_to,
        quiet=bool(args.quiet),
        dry_run=bool(args.dry_run),
    )
    if args.dry_run:
        print(f"[dry-run] 未写入文件；若导出请去掉 --dry-run，默认输出：{out}")
    else:
        print(f"完成：{out}")


if __name__ == "__main__":
    main()

