# -*- coding: utf-8 -*-
"""
全表一致性自检 + 自动修正

- 读取 `不同剂量组ADR分析 (TFL).xlsx`（当前工作簿 active sheet）
- 读取 `不同剂量组ADR分析-分级 (TFL).pdf` 并解析首选术语的 Total/1级/2级/3级（跨页、断表）
- 对比 Excel 每个 ADR 块（A列术语 + B列分级 1级/2级/3级/Total）与 PDF 解析值
- 发现不一致则写回修正，并生成差异报告 `consistency_report.txt`
"""

from __future__ import annotations

import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from fill_adr_from_pdf import (
    PDF_GRADE,
    XLSX_PATH,
    collect_graded_adr_data,
    find_pdf_term_for_excel,
)


BASE = Path(__file__).resolve().parent
REPORT_PATH = BASE / "consistency_report.txt"


GRADE_ROWS = ("1级", "2级", "3级", "Total")

# Excel: 5组 * 3列，从 C 到 Q
START_COL = 3
GROUPS = 5
CELLS_PER_GROUP = 3


def _norm_str(v: Any) -> str:
    return str(v or "").strip()


def _is_close(a: Any, b: Any, tol: float = 1e-6) -> bool:
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    # int/float 比较
    try:
        fa = float(a)
        fb = float(b)
        if math.isfinite(fa) and math.isfinite(fb):
            return abs(fa - fb) <= tol
    except Exception:
        pass
    return _norm_str(a) == _norm_str(b)


def _expected_cells(groups_triplets: list[tuple[int, int, float]]) -> list[Any]:
    """
    groups_triplets: [(例次, 例数n, rate), ...] from parse_row_to_groups
    Excel列顺序：例数(n)、例次、发生率
    """
    out: list[Any] = []
    for ex_count, n_subj, rate in groups_triplets:
        out.extend([n_subj, ex_count, rate])
    return out


def audit_and_fix() -> tuple[int, int]:
    """
    Returns: (mismatch_count, fixed_cell_count)
    """
    adr_data = collect_graded_adr_data(PDF_GRADE)

    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    mismatches = 0
    fixed_cells = 0
    lines: list[str] = []

    row = 1
    while row <= ws.max_row:
        grade = _norm_str(ws.cell(row, 2).value)
        if grade != "1级":
            row += 1
            continue

        adr_name = _norm_str(ws.cell(row, 1).value)
        if not adr_name:
            row += 1
            continue

        pdf_term = find_pdf_term_for_excel(adr_name)
        if not pdf_term or pdf_term not in adr_data:
            lines.append(f"[SKIP] {adr_name}: PDF中未找到术语")
            row += 4
            continue

        for ri, gr in enumerate(GRADE_ROWS):
            r = row + ri
            if r > ws.max_row:
                break
            if _norm_str(ws.cell(r, 2).value) != gr:
                continue
            if gr not in adr_data[pdf_term]:
                lines.append(f"[SKIP] {adr_name} {gr}: PDF中无该分级行")
                continue

            expected = _expected_cells(adr_data[pdf_term][gr])
            # 读 Excel 当前行 C..Q
            current = [ws.cell(r, c).value for c in range(START_COL, START_COL + GROUPS * CELLS_PER_GROUP)]

            # 对比并修正
            row_mismatch = False
            for idx, (cur, exp) in enumerate(zip(current, expected)):
                if not _is_close(cur, exp):
                    row_mismatch = True
                    col = START_COL + idx
                    ws.cell(r, col, value=exp)
                    fixed_cells += 1
            if row_mismatch:
                mismatches += 1
                lines.append(f"[FIX] {adr_name} {gr}: 已修正不一致单元格")

        row += 4

    wb.save(XLSX_PATH)

    REPORT_PATH.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8")
    return mismatches, fixed_cells


if __name__ == "__main__":
    mismatches, fixed_cells = audit_and_fix()
    print(f"done. mismatches={mismatches}, fixed_cells={fixed_cells}, report={REPORT_PATH}")

