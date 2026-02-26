# -*- coding: utf-8 -*-
"""
参照 `不同剂量组ADR分析-Template.xlsx` 的制图逻辑/格式/颜色风格，
在 `不同剂量组ADR分析 (TFL).xlsx` 中生成同款图表，覆盖 5 个组：
低剂量试验组、高剂量试验组、低剂量佐剂组、高剂量佐剂组、安慰剂组。

模板图表（经 openpyxl 读取）逻辑：
- 柱状图 BarChart (type='col')
- categories: ADR 名称（A3,A7,A11,...）
- values: 各组 Total 行“例数(n)”列（C6,F6,I6,L6,O6 等；每 4 行一个 ADR，Total 行在 6,10,14...）
- 图例在底部，图表放置位置使用模板 anchor

说明：
openpyxl 的 Reference 不支持 step，因此用 Excel tuple 公式：
(Sheet!$C$6,Sheet!$C$10,...) 来引用非连续行，和模板一致。
"""

from __future__ import annotations

from copy import deepcopy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart
from openpyxl.chart.data_source import (
    AxDataSource,
    NumDataSource,
    NumRef,
    StrRef,
)
from openpyxl.chart.series import Series, SeriesLabel
from src.color_theme import get_series_color


BASE = Path(__file__).resolve().parent
TEMPLATE_PATTERN = "*Template.xlsx"
TFL = BASE / "不同剂量组ADR分析 (TFL).xlsx"


def get_template_path() -> Path:
    matches = list(BASE.glob(TEMPLATE_PATTERN))
    if not matches:
        raise FileNotFoundError(
            "未找到模板文件（*Template.xlsx）。请将模板放在 01_Excel_Charts 目录下。"
        )
    return matches[0]


def _total_rows(max_row: int) -> list[int]:
    return list(range(6, max_row + 1, 4))


def _adr_rows(max_row: int) -> list[int]:
    return list(range(3, max_row + 1, 4))


def _pick_anchor_and_style(tws):
    if not tws._charts:
        return None, None
    ch = tws._charts[0]
    return deepcopy(ch.anchor), ch


def tuple_formula(sheet_title: str, col_letter: str, rows: list[int]) -> str:
    parts = [f"'{sheet_title}'!${col_letter}${r}" for r in rows]
    return f"({','.join(parts)})"


def apply():
    template_path = get_template_path()
    twb = load_workbook(template_path)
    tws = twb.active
    anchor, tmpl_chart = _pick_anchor_and_style(tws)

    wb = load_workbook(TFL)
    ws = wb.active

    # 清理已有图表，避免重复叠加
    ws._charts = []

    adr_rows = _adr_rows(ws.max_row)
    total_rows = _total_rows(ws.max_row)
    cats_formula = tuple_formula(ws.title, "A", adr_rows)

    group_cols = [
        ("C", "低剂量试验组", get_series_color("低剂量试验组")),
        ("F", "高剂量试验组", get_series_color("高剂量试验组")),
        ("I", "低剂量佐剂组", get_series_color("低剂量佐剂组")),
        ("L", "高剂量佐剂组", get_series_color("高剂量佐剂组")),
        ("O", "安慰剂组", get_series_color("安慰剂组")),
    ]

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.overlap = 0
    chart.title = "不同剂量组ADR发生情况"
    chart.legend.position = "b"
    chart.legend.overlay = False
    chart.y_axis.title = ""
    chart.x_axis.title = ""

    for col_letter, name, color in group_cols:
        values_formula = tuple_formula(ws.title, col_letter, total_rows)
        s = Series()
        s.tx = SeriesLabel(v=name)
        s.val = NumDataSource(numRef=NumRef(f=values_formula))
        s.cat = AxDataSource(strRef=StrRef(f=cats_formula))
        s.graphicalProperties.solidFill = color
        chart.series.append(s)

    # 尽量复制模板能复制的样式字段
    if tmpl_chart is not None:
        for attr in ("style", "gapWidth", "dLbls", "dataLabels"):
            if hasattr(tmpl_chart, attr):
                try:
                    setattr(chart, attr, deepcopy(getattr(tmpl_chart, attr)))
                except Exception:
                    pass

    if anchor is not None:
        chart.anchor = anchor
        ws.add_chart(chart)
    else:
        ws.add_chart(chart, "L6")

    wb.save(TFL)
    print("done:", TFL)


if __name__ == "__main__":
    apply()

