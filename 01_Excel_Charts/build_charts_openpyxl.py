"""
在已经“无图形部件”的 `不同剂量组ADR分析 (TFL).xlsx` 上，
用 openpyxl 重新创建与 Template 风格一致的组合图（柱形 + 折线）。

设计：
- 新建隐藏 sheet `ChartData`，整理：
  A列：ADR 名称（A3,7,11,...）
  B–F列：各组 Total 行“例数(n)”  (C/F/I/L/O)
  G–K列：各组 Total 行“发生率”    (E/H/K/N/Q)
- 用连续区域创建：
  - 柱图：BarChart，values = ChartData!B2:F(n)，cats = A2:A(n)
  - 折线：LineChart，values = ChartData!G2:K(n)，共用 categories
  - 组合：bar += line
"""

from __future__ import annotations

from copy import deepcopy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference


BASE = Path(__file__).resolve().parent
TFL = BASE / "不同剂量组ADR分析 (TFL).xlsx"
TEMPLATE = BASE / "不同剂量组ADR分析-Template.xlsx"


def _adr_total_rows(ws):
    """返回 (adr_rows, total_rows) 列表。"""
    adr_rows = []
    total_rows = []
    r = 3
    while r <= ws.max_row:
        if ws.cell(r, 1).value:
            adr_rows.append(r)
        if ws.cell(r + 3, 2).value == "Total":
            total_rows.append(r + 3)
        r += 4
    return adr_rows, total_rows


def build():
    wb = load_workbook(TFL)
    ws = wb.active

    # 删除旧的 ChartData
    if "ChartData" in wb.sheetnames:
        del wb["ChartData"]
    cds = wb.create_sheet("ChartData")

    adr_rows, total_rows = _adr_total_rows(ws)
    sheet_name = ws.title
    safe_sheet = sheet_name.replace("'", "''")
    sheet_ref = f"'{safe_sheet}'"

    # 头部
    cds["A1"] = "ADR"
    cds["B1"] = "低剂量试验组-例数"
    cds["C1"] = "高剂量试验组-例数"
    cds["D1"] = "低剂量佐剂组-例数"
    cds["E1"] = "高剂量佐剂组-例数"
    cds["F1"] = "安慰剂组-例数"

    cds["G1"] = "低剂量试验组-发生率"
    cds["H1"] = "高剂量试验组-发生率"
    cds["I1"] = "低剂量佐剂组-发生率"
    cds["J1"] = "高剂量佐剂组-发生率"
    cds["K1"] = "安慰剂组-发生率"

    # 填充数据（使用公式 link 回原表，而不是静态数值）
    for idx, (ar, tr) in enumerate(zip(adr_rows, total_rows), start=2):
        # ADR 名称直接引用原表 A 行
        cds[f"A{idx}"] = f"={sheet_ref}!A{ar}"
        # 例数(n)：C,F,I,L,O
        cds[f"B{idx}"] = f"={sheet_ref}!C{tr}"
        cds[f"C{idx}"] = f"={sheet_ref}!F{tr}"
        cds[f"D{idx}"] = f"={sheet_ref}!I{tr}"
        cds[f"E{idx}"] = f"={sheet_ref}!L{tr}"
        cds[f"F{idx}"] = f"={sheet_ref}!O{tr}"
        # 发生率：E,H,K,N,Q
        cds[f"G{idx}"] = f"={sheet_ref}!E{tr}"
        cds[f"H{idx}"] = f"={sheet_ref}!H{tr}"
        cds[f"I{idx}"] = f"={sheet_ref}!K{tr}"
        cds[f"J{idx}"] = f"={sheet_ref}!N{tr}"
        cds[f"K{idx}"] = f"={sheet_ref}!Q{tr}"

    max_row = len(adr_rows) + 1

    # 构建柱状图
    data_bar = Reference(cds, min_col=2, min_row=1, max_col=6, max_row=max_row)
    cats = Reference(cds, min_col=1, min_row=2, max_row=max_row)

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.overlap = 0
    bar.title = "不同剂量组ADR发生情况"
    bar.y_axis.title = "例数"
    bar.legend.position = "b"
    bar.legend.overlay = False
    bar.add_data(data_bar, titles_from_data=True)
    bar.set_categories(cats)

    # 颜色/样式尽量按模板拷贝
    if TEMPLATE.exists():
        twb = load_workbook(TEMPLATE)
        tws = twb.active
        if tws._charts:
            tmpl = tws._charts[0]
            if hasattr(tmpl, "style"):
                bar.style = tmpl.style
            # 复制图例和网格线风格（模板背景无网格线）
            try:
                bar.legend = deepcopy(tmpl.legend)
            except Exception:
                pass

    # 去掉主坐标轴网格线
    bar.y_axis.majorGridlines = None

    # 构建折线图（发生率）
    data_line = Reference(cds, min_col=7, min_row=1, max_col=11, max_row=max_row)
    line = LineChart()
    line.y_axis.axId = 200
    line.y_axis.title = "发生率"
    line.add_data(data_line, titles_from_data=True)
    line.set_categories(cats)
    line.y_axis.crosses = "max"
    # 使用直线（非平滑曲线），并去掉次轴网格线
    for s in line.series:
        s.smooth = False
    line.y_axis.majorGridlines = None

    # 组合
    bar += line

    # 清除原 sheet 上的图表（如果 strip 已干净则为空）
    ws._charts = []
    ws.add_chart(bar, "L6")

    # 可选：隐藏 ChartData
    cds.sheet_state = "hidden"

    wb.save(TFL)
    print("charts rebuilt on", TFL)


if __name__ == "__main__":
    build()

