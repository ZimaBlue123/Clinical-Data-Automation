"""
fill_clinical_table.py - 临床试验表格数据填充（统一入口）

功能：
  统一处理GMC（几何平均浓度）、GMI（几何平均倍数）和阳转率三类临床试验
  数据表格的自动填充。支持任意子表（含中文/特殊字符），自动识别源数据位置。

支持表格类型：
  - gmc:       GMC（几何平均浓度）表格
  - gmi:       GMI（几何平均倍数）表格
  - yangzhuai: 阳转率表格
  - all:       处理所有支持的子表（默认）

用法：
    python fill_clinical_table.py <excel_file> [选项]

参数：
    excel_file                    Excel文件路径（必填）

选项：
    --type {gmc,gmi,yangzhuai,all}
                                  表格类型（默认: all 自动检测）
    --sheets SHEETS               指定子表（逗号分隔，优先级高于--type）
    --output-dir DIR              输出目录 (默认: 同级output目录)
    --output-name NAME            输出文件名 (默认: 输入文件名)
    --no-include-pre              GMC表格不填充免前(第3行)
    -v, --verbose                 显示详细日志
    -h, --help                    显示帮助

示例：
    # 处理所有支持的子表（自动检测GMC/GMI/阳转率）
    python fill_clinical_table.py input/TVAX-006.xlsx

    # 仅处理GMC子表
    python fill_clinical_table.py input/TVAX-006.xlsx --type gmc

    # 仅处理GMI子表
    python fill_clinical_table.py input/TVAX-006.xlsx --type gmi

    # 仅处理阳转率子表
    python fill_clinical_table.py input/TVAX-006.xlsx --type yangzhuai

    # 指定子表
    python fill_clinical_table.py input/TVAX-006.xlsx --sheets "总体GMC,40-59岁GMC"

    # 自定义输出
    python fill_clinical_table.py input/TVAX-006.xlsx --output-dir ./output

GMC表格数据格式：
    GMC表格结构（每子表）：
      第1行：组别标题
      第2行：子标题（平均值、上限、下限）
      第3行：免前（GMC值）
      第4-7行：4个时间点

    源数据：
      行12: GMC (95%CI) - 用于第3行免前
      行17/22/27/32: LS GMC (95%CI) - 用于第4-7行

    列结构：
      B-D  (列2-4):  低剂量佐剂组
      E-G  (列5-7):  高剂量佐剂组
      H-J  (列8-10): 低剂量试验组
      K-M  (列11-13): 高剂量试验组
      N-P  (列14-16): 安慰剂组

GMI表格数据格式：
    GMI表格结构（每子表）：同GMC。
    源数据：脚本动态扫描定位"GMI (95%CI)"行（适配不同子表行号差异）。
    源数据格式：同GMC（如 "1.12 (0.91, 1.38)"）。

阳转率表格数据格式：
    阳转率表格结构（每子表）：同GMC，但第3行免前不填。
    源数据：脚本自动扫描定位
      - 标题行：A列含"一免后"或"全免后"
      - "阳转例数（阳转率）"行：格式 "24 (75.00)"
      - "95%CI"行：格式 "56.60, 88.54"
"""

from __future__ import annotations

import argparse
import logging
import os
import re
import sys
from pathlib import Path
from typing import Any
from collections.abc import Iterable

# ====== 通用配置 ======

# 详细统计表中的5个组的源列（C-G，即列3-7）
SOURCE_GROUP_COLS: list[int] = [3, 4, 5, 6, 7]

# 表格中的目标起始列（B,E,H,K,N）
TARGET_GROUP_BASE_COLS: list[int] = [2, 5, 8, 11, 14]

# ====== GMC 配置 ======
GMC_TIMEPOINTS: dict[int, str] = {
    3: '免前',
    4: '一免后1个月',
    5: '一免后2个月',
    6: '全免后1个月',
    7: '全免后6个月',
}

# GMC源行映射：目标行 -> 源行
GMC_SOURCE_ROW_MAP: dict[int, int] = {
    3: 12,   # 免前 ← 行12 (GMC)
    4: 17,   # 一免后1个月 ← 行17 (LS GMC)
    5: 22,
    6: 27,
    7: 32,
}

# GMC子表关键字
GMC_SHEET_KEYWORDS: list[str] = ['GMC']

# ====== 阳转率 配置 ======
YANGZHUAI_TIMEPOINTS: dict[int, str] = {
    4: '一免后1个月',
    5: '一免后2个月',
    6: '全免后1个月',
    7: '全免后6个月',
}

# 阳转率子表关键字
YANGZHUAI_SHEET_KEYWORDS: list[str] = ['阳转率']

# ====== GMI 配置 ======
GMI_TIMEPOINTS: dict[int, str] = {
    4: '一免后1个月',
    5: '一免后2个月',
    6: '全免后1个月',
    7: '全免后6个月',
}

# GMI源行映射：默认源行（脚本动态扫描优先）
GMI_SOURCE_ROW_MAP: dict[int, int] = {
    4: 18,
    5: 29,
    6: 41,
    7: 52,
}

# GMI子表关键字
GMI_SHEET_KEYWORDS: list[str] = ['GMI']

# 子表类型常量
TABLE_TYPE_GMC = 'gmc'
TABLE_TYPE_GMI = 'gmi'
TABLE_TYPE_YANGZHUAI = 'yangzhuai'
TABLE_TYPE_ALL = 'all'

# ====== 日志配置 ======

logger = logging.getLogger('fill_clinical_table')


def _setup_logging(verbose: bool = False) -> None:
    """配置日志"""
    level = logging.DEBUG if verbose else logging.INFO
    handler = logging.StreamHandler(sys.stdout)
    handler.setFormatter(logging.Formatter(
        fmt='%(message)s',
    ))
    logger.handlers.clear()
    logger.addHandler(handler)
    logger.setLevel(level)


# ====== 通用工具函数 ======

def _to_str(v: Any) -> str:
    """安全转换为字符串"""
    if v is None:
        return ''
    return str(v).strip()


def _to_float(v: Any) -> float | None:
    """安全转换为float"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip())
        except (ValueError, AttributeError):
            return None
    return None


def parse_gmc_ci(gmc_str: Any) -> tuple[float | None, float | None, float | None]:
    """
    解析GMC (95%CI)格式字符串，支持：
      - "768.17(507.87, 1161.89)"
      - "644.46 (280.78, 1479.20)"

    Returns:
        tuple: (平均值, 上限, 下限) 或 (None, None, None)
    """
    if gmc_str is None:
        return None, None, None
    if isinstance(gmc_str, (int, float)):
        return float(gmc_str), None, None
    if not isinstance(gmc_str, str):
        return None, None, None

    s = gmc_str.strip()
    if not s:
        return None, None, None

    pattern = r'([\d.]+)\s*\(\s*([\d.]+)\s*,\s*([\d.]+)\s*\)'
    match = re.search(pattern, s)
    if match:
        try:
            mean = float(match.group(1))
            lower = float(match.group(2))
            upper = float(match.group(3))
            return mean, upper, lower
        except (ValueError, IndexError):
            return None, None, None

    return None, None, None


def parse_positive_rate(rate_str: Any) -> float | None:
    """
    解析阳转率字符串："24 (75.00)" → 75.0
    """
    if rate_str is None:
        return None
    if isinstance(rate_str, (int, float)):
        return float(rate_str)
    if not isinstance(rate_str, str):
        return None

    s = rate_str.strip()
    if not s:
        return None

    pattern = r'\d+\s*\(\s*([\d.]+)\s*\)'
    match = re.search(pattern, s)
    if match:
        try:
            return float(match.group(1))
        except ValueError:
            return None

    if re.match(r'^\d+(\.\d+)?$', s):
        try:
            return float(s)
        except ValueError:
            return None

    return None


def parse_ci(ci_str: Any) -> tuple[float | None, float | None]:
    """
    解析95%CI字符串："56.60, 88.54" → (56.60, 88.54)
    """
    if ci_str is None:
        return None, None
    if not isinstance(ci_str, str):
        return None, None

    s = ci_str.strip()
    if not s:
        return None, None

    pattern = r'([\d.]+)\s*,\s*([\d.]+)'
    match = re.search(pattern, s)
    if match:
        try:
            return float(match.group(1)), float(match.group(2))
        except ValueError:
            return None, None

    return None, None


def detect_sheet_type(sheet_name: str) -> str:
    """
    检测子表类型

    Returns:
        str: 'gmc' | 'gmi' | 'yangzhuai' | 'unknown'
    """
    if not isinstance(sheet_name, str):
        return 'unknown'
    # GMI需优先于GMC检查（避免"GMI"被"GMC"匹配）
    for kw in GMI_SHEET_KEYWORDS:
        if kw in sheet_name:
            return TABLE_TYPE_GMI
    for kw in GMC_SHEET_KEYWORDS:
        if kw in sheet_name:
            return TABLE_TYPE_GMC
    for kw in YANGZHUAI_SHEET_KEYWORDS:
        if kw in sheet_name:
            return TABLE_TYPE_YANGZHUAI
    return 'unknown'


def _safe_sheet_get(ws, row: int, col: int) -> Any:
    """安全获取单元格值（边界检查）"""
    try:
        if row < 1 or col < 1:
            return None
        if row > ws.max_row or col > ws.max_column:
            return None
        return ws.cell(row=row, column=col).value
    except (IndexError, AttributeError) as e:
        logger.debug(f"安全获取单元格失败 row={row}, col={col}: {e}")
        return None


# ====== GMC 填充逻辑 ======

def fill_gmc_sheet(ws, sheet_name: str, include_pre: bool = True, verbose: bool = False) -> int:
    """
    填充GMC子表

    Args:
        include_pre: 是否填充第3行（免前）

    Returns:
        int: 成功填充的单元格数量
    """
    if verbose:
        logger.info(f"\n{'=' * 80}")
        logger.info(f"[GMC] 处理子表: {sheet_name} (行数: {ws.max_row}, 列数: {ws.max_column})")
        logger.info('=' * 80)

    filled_count = 0

    for tgt_row, timepoint_name in GMC_TIMEPOINTS.items():
        if tgt_row == 3 and not include_pre:
            if verbose:
                logger.info("\n跳过 免前 (第3行, include_pre=False)")
            continue

        src_row = GMC_SOURCE_ROW_MAP.get(tgt_row)
        if not src_row:
            continue

        # 边界检查
        if src_row > ws.max_row:
            if verbose:
                logger.warning(f"\n跳过 {timepoint_name} (目标行{tgt_row}): 源行{src_row}不存在于工作表(最大行={ws.max_row})")
            continue

        if verbose:
            logger.info(f"\n填充 {timepoint_name} (目标行{tgt_row}, 源行{src_row}):")

        for i, src_col in enumerate(SOURCE_GROUP_COLS):
            cell_value = _safe_sheet_get(ws, src_row, src_col)
            mean, upper, lower = parse_gmc_ci(cell_value)

            if mean is not None and upper is not None and lower is not None:
                tgt_base = TARGET_GROUP_BASE_COLS[i]
                ws.cell(row=tgt_row, column=tgt_base).value = mean
                ws.cell(row=tgt_row, column=tgt_base + 1).value = upper
                ws.cell(row=tgt_row, column=tgt_base + 2).value = lower
                if verbose:
                    logger.info(f"  组{i + 1}: '{cell_value}' -> 均={mean}, 上={upper}, 下={lower}")
                filled_count += 3
            elif verbose:
                logger.info(f"  组{i + 1}: '{cell_value}' -> 无效或缺失，跳过")

    return filled_count


# ====== 阳转率 填充逻辑 ======

def _find_timepoint_blocks_yangzhuai(ws) -> dict[int, dict[str, Any]]:
    """
    扫描阳转率子表，自动定位各时间点的源数据行
    """
    timepoint_blocks: dict[int, dict[str, Any]] = {}

    title_rows: list[tuple[int, str, str]] = []
    for row_idx in range(8, ws.max_row + 1):
        first_col = _to_str(_safe_sheet_get(ws, row_idx, 1))
        if not first_col:
            continue
        for tp_name in YANGZHUAI_TIMEPOINTS.values():
            if tp_name in first_col:
                title_rows.append((row_idx, tp_name, first_col))
                break

    for tgt_row, tp_name in YANGZHUAI_TIMEPOINTS.items():
        title_row = None
        for tr, tn, _ in title_rows:
            if tn == tp_name:
                title_row = tr
                break

        if title_row is None:
            continue

        rate_row = None
        ci_row = None
        for r in range(title_row, min(title_row + 6, ws.max_row + 1)):
            second_col = _to_str(_safe_sheet_get(ws, r, 2))
            if not second_col:
                continue
            if '阳转例数' in second_col:
                rate_row = r
            if '95%CI' in second_col or '95% CI' in second_col:
                ci_row = r

        if rate_row and ci_row:
            timepoint_blocks[tgt_row] = {
                'rate_row': rate_row,
                'ci_row': ci_row,
                'title': tp_name,
            }

    return timepoint_blocks


def fill_yangzhuai_sheet(ws, sheet_name: str, verbose: bool = False) -> int:
    """
    填充阳转率子表
    """
    if verbose:
        logger.info(f"\n{'=' * 80}")
        logger.info(f"[阳转率] 处理子表: {sheet_name} (行数: {ws.max_row}, 列数: {ws.max_column})")
        logger.info('=' * 80)

    timepoint_blocks = _find_timepoint_blocks_yangzhuai(ws)

    if not timepoint_blocks:
        if verbose:
            logger.warning("  未找到时间点源数据")
        return 0

    if verbose:
        logger.info("\n识别到的时间点:")
        for tgt, info in sorted(timepoint_blocks.items()):
            logger.info(f"  行{tgt} {info['title']}: 阳转率行{info['rate_row']}, CI行{info['ci_row']}")

    filled_count = 0

    for tgt_row, info in sorted(timepoint_blocks.items()):
        rate_row = info['rate_row']
        ci_row = info['ci_row']
        title = info['title']

        if verbose:
            logger.info(f"\n填充 {title} (目标行{tgt_row}):")

        for i, src_col in enumerate(SOURCE_GROUP_COLS):
            rate_val = parse_positive_rate(_safe_sheet_get(ws, rate_row, src_col))
            ci_lower, ci_upper = parse_ci(_safe_sheet_get(ws, ci_row, src_col))

            if rate_val is not None and ci_lower is not None and ci_upper is not None:
                tgt_base = TARGET_GROUP_BASE_COLS[i]
                ws.cell(row=tgt_row, column=tgt_base).value = rate_val
                ws.cell(row=tgt_row, column=tgt_base + 1).value = ci_upper
                ws.cell(row=tgt_row, column=tgt_base + 2).value = ci_lower
                if verbose:
                    logger.info(f"  组{i + 1}: 率={rate_val}%, CI=({ci_lower}, {ci_upper})")
                filled_count += 3
            elif verbose:
                rate_str = _safe_sheet_get(ws, rate_row, src_col)
                ci_str = _safe_sheet_get(ws, ci_row, src_col)
                logger.info(f"  组{i + 1}: 阳转率='{rate_str}', CI='{ci_str}' -> 数据无效，跳过")

    return filled_count


# ====== GMI 填充逻辑 ======

def _find_gmi_source_rows(ws, verbose: bool = False) -> dict[int, int]:
    """
    动态扫描GMI子表，查找各时间点对应的GMI (95%CI)源行
    保证不同行号结构（如"60岁以上GMI"是51行）的子表都能正确填充

    Returns:
        dict: {目标行号: 源行号}
    """
    title_rows: list[tuple[int, str, str]] = []
    for r in range(8, ws.max_row + 1):
        first_col = _to_str(_safe_sheet_get(ws, r, 1))
        if not first_col:
            continue
        for tp_name in GMI_TIMEPOINTS.values():
            if tp_name in first_col:
                title_rows.append((r, tp_name, first_col))
                break

    result: dict[int, int] = {}
    for tgt_row, tp_name in GMI_TIMEPOINTS.items():
        title_row = None
        for tr, tn, _ in title_rows:
            if tn == tp_name:
                title_row = tr
                break
        if title_row is None:
            continue

        gmi_row = None
        for r in range(title_row, min(title_row + 12, ws.max_row + 1)):
            second_col = _to_str(_safe_sheet_get(ws, r, 2))
            if 'GMI (95%CI)' in second_col or 'GMI(95%CI)' in second_col:
                gmi_row = r
                break

        if gmi_row:
            result[tgt_row] = gmi_row

    return result


def fill_gmi_sheet(ws, sheet_name: str, verbose: bool = False) -> int:
    """
    填充GMI子表
    复用parse_gmc_ci解析器（格式与GMC一致）
    动态扫描源行号以适应不同子表结构
    """
    if verbose:
        logger.info(f"\n{'=' * 80}")
        logger.info(f"[GMI] 处理子表: {sheet_name} (行数: {ws.max_row}, 列数: {ws.max_column})")
        logger.info('=' * 80)

    source_rows = _find_gmi_source_rows(ws, verbose)

    if not source_rows:
        if verbose:
            logger.warning("  未找到任何GMI (95%CI)源行")
        return 0

    if verbose:
        logger.info("\n动态识别的源行号:")
        for tgt_row in sorted(source_rows):
            tp_name = GMI_TIMEPOINTS.get(tgt_row, f'行{tgt_row}')
            logger.info(f"  行{tgt_row} {tp_name} <- 源行{source_rows[tgt_row]}")

    filled_count = 0

    for tgt_row, timepoint_name in GMI_TIMEPOINTS.items():
        if tgt_row not in source_rows:
            if verbose:
                logger.info(f"\n跳过 {timepoint_name} (目标行{tgt_row}): 未找到源行")
            continue

        src_row = source_rows[tgt_row]

        if verbose:
            logger.info(f"\n填充 {timepoint_name} (目标行{tgt_row}, 源行{src_row}):")

        for i, src_col in enumerate(SOURCE_GROUP_COLS):
            cell_value = _safe_sheet_get(ws, src_row, src_col)
            mean, upper, lower = parse_gmc_ci(cell_value)

            if mean is not None and upper is not None and lower is not None:
                tgt_base = TARGET_GROUP_BASE_COLS[i]
                ws.cell(row=tgt_row, column=tgt_base).value = mean
                ws.cell(row=tgt_row, column=tgt_base + 1).value = upper
                ws.cell(row=tgt_row, column=tgt_base + 2).value = lower
                if verbose:
                    logger.info(f"  组{i + 1}: '{cell_value}' -> 均={mean}, 上={upper}, 下={lower}")
                filled_count += 3
            elif verbose:
                logger.info(f"  组{i + 1}: '{cell_value}' -> 无效或缺失，跳过")

    return filled_count


# ====== 工作簿处理 ======

def process_workbook(
    excel_path: str,
    sheet_tasks: list[tuple[str, str]],
    output_path: str,
    verbose: bool = False,
) -> dict[str, Any] | None:
    """
    处理工作簿

    Args:
        sheet_tasks: [(sheet_name, table_type), ...]

    Returns:
        dict: 处理结果统计，失败时返回None
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("需要安装openpyxl库: pip install openpyxl")
        return None

    if not os.path.exists(excel_path):
        logger.error(f"输入文件不存在: {excel_path}")
        return None

    if not os.access(excel_path, os.R_OK):
        logger.error(f"输入文件不可读: {excel_path}")
        return None

    logger.info(f"读取文件: {excel_path}")

    try:
        wb = openpyxl.load_workbook(excel_path)
    except (openpyxl.utils.exceptions.InvalidFileException, Exception) as e:
        logger.error(f"加载Excel文件失败: {e}")
        return None

    # 验证子表
    missing = [s for s, _ in sheet_tasks if s not in wb.sheetnames]
    if missing:
        logger.warning(f"以下子表不存在: {missing}")
        logger.warning(f"可用子表: {wb.sheetnames}")
        sheet_tasks = [(s, t) for s, t in sheet_tasks if s in wb.sheetnames]
        if not sheet_tasks:
            logger.error("没有可处理的子表")
            return None

    type_groups: dict[str, list[str]] = {
        TABLE_TYPE_GMC: [],
        TABLE_TYPE_GMI: [],
        TABLE_TYPE_YANGZHUAI: [],
    }
    for sheet_name, table_type in sheet_tasks:
        if table_type in type_groups:
            type_groups[table_type].append(sheet_name)

    total_filled = 0
    sheet_results: dict[str, int] = {}

    for table_type, sheets in type_groups.items():
        if not sheets:
            continue
        logger.info(f"\n[{table_type.upper()}] 处理 {len(sheets)} 个子表: {sheets}")
        for sheet_name in sheets:
            try:
                ws = wb[sheet_name]
            except KeyError:
                logger.warning(f"  跳过 {sheet_name}: 子表不存在")
                continue

            try:
                if table_type == TABLE_TYPE_GMC:
                    filled = fill_gmc_sheet(ws, sheet_name, include_pre=True, verbose=verbose)
                elif table_type == TABLE_TYPE_GMI:
                    filled = fill_gmi_sheet(ws, sheet_name, verbose=verbose)
                elif table_type == TABLE_TYPE_YANGZHUAI:
                    filled = fill_yangzhuai_sheet(ws, sheet_name, verbose=verbose)
                else:
                    continue
                sheet_results[sheet_name] = filled
                total_filled += filled
                if not verbose:
                    logger.info(f"  [OK] {sheet_name}: 填充 {filled} 个单元格")
            except Exception as e:
                logger.error(f"  处理 {sheet_name} 时出错: {e}", exc_info=verbose)
                sheet_results[sheet_name] = 0

    # 保存
    output_dir = os.path.dirname(output_path)
    if output_dir:
        try:
            os.makedirs(output_dir, exist_ok=True)
        except (OSError, PermissionError) as e:
            logger.error(f"创建输出目录失败: {e}")
            return None

    try:
        wb.save(output_path)
    except (PermissionError, OSError) as e:
        logger.error(f"保存文件失败: {e}")
        return None

    logger.info(f"\n文件已保存: {output_path}")
    logger.info(f"\n总计: 处理 {len(sheet_tasks)} 个子表，填充 {total_filled} 个单元格")

    return {
        'sheets': sheet_results,
        'total_filled': total_filled,
        'output_path': output_path,
    }


def resolve_output_path(
    excel_path: str,
    output_dir: str | None = None,
    output_name: str | None = None,
) -> Path:
    """解析输出路径"""
    excel_path_obj = Path(excel_path).resolve()
    input_dir = excel_path_obj.parent
    input_stem = excel_path_obj.stem
    input_suffix = excel_path_obj.suffix

    if output_dir is None:
        output_dir = str(input_dir / 'output')

    output_dir_obj = Path(output_dir)

    if output_name is None:
        output_name = f"{input_stem}{input_suffix}"
    else:
        output_name_obj = Path(output_name)
        if not output_name_obj.suffix:
            output_name = f"{output_name}{input_suffix}"

    return output_dir_obj / output_name


def build_sheet_tasks(
    wb,
    table_type: str,
    explicit_sheets: Iterable[str] | None = None,
) -> list[tuple[str, str]]:
    """
    构建要处理的子表任务列表

    Args:
        table_type: 'gmc' | 'gmi' | 'yangzhuai' | 'all'
        explicit_sheets: 用户显式指定的子表列表 (None=自动检测)
    """
    sheet_tasks: list[tuple[str, str]] = []

    if explicit_sheets:
        for s in explicit_sheets:
            t = detect_sheet_type(s)
            if t == 'unknown':
                logger.warning(f"无法识别子表类型 '{s}'，跳过")
                continue
            if table_type != TABLE_TYPE_ALL and t != table_type:
                logger.warning(f"子表 '{s}' (类型={t}) 与 --type={table_type} 不匹配，跳过")
                continue
            sheet_tasks.append((s, t))
    else:
        for s in wb.sheetnames:
            t = detect_sheet_type(s)
            if t == 'unknown':
                continue
            if table_type != TABLE_TYPE_ALL and t != table_type:
                continue
            sheet_tasks.append((s, t))

    return sheet_tasks


def main(argv: list[str] | None = None) -> int:
    """主入口"""
    parser = argparse.ArgumentParser(
        description='临床试验表格数据填充（统一入口：GMC + GMI + 阳转率）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument('excel_file', nargs='?', help='Excel文件路径')
    parser.add_argument(
        '--type', '-t',
        choices=[TABLE_TYPE_GMC, TABLE_TYPE_GMI, TABLE_TYPE_YANGZHUAI, TABLE_TYPE_ALL],
        default=TABLE_TYPE_ALL,
        help='表格类型 (默认: all 自动检测所有)',
    )
    parser.add_argument(
        '--sheets', '-s',
        help='指定子表（逗号分隔），优先级高于--type',
    )
    parser.add_argument(
        '--output-dir', '-o',
        help='输出目录 (默认: 同级output目录)',
    )
    parser.add_argument(
        '--output-name', '-n',
        help='输出文件名 (默认: 输入文件名)',
    )
    parser.add_argument(
        '--no-include-pre',
        action='store_true',
        help='GMC表格不填充免前(第3行)',
    )
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='显示详细日志',
    )

    args = parser.parse_args(argv)
    _setup_logging(args.verbose)

    if not args.excel_file:
        parser.print_help()
        logger.error("请提供Excel文件路径")
        return 1

    excel_path = os.path.abspath(args.excel_file)

    if not os.path.exists(excel_path):
        logger.error(f"文件不存在: {excel_path}")
        return 1

    if not os.access(excel_path, os.R_OK):
        logger.error(f"文件不可读: {excel_path}")
        return 1

    try:
        import openpyxl
    except ImportError:
        logger.error("需要安装openpyxl库: pip install openpyxl")
        return 1

    # 预扫描以构建任务列表
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
    except Exception as e:
        logger.error(f"无法读取Excel文件: {e}")
        return 1

    explicit = [s.strip() for s in args.sheets.split(',')] if args.sheets else None
    sheet_tasks = build_sheet_tasks(wb, args.type, explicit)
    wb.close()

    if not sheet_tasks:
        logger.error("没有匹配条件的子表")
        return 1

    # 解析输出路径
    try:
        output_path = resolve_output_path(excel_path, args.output_dir, args.output_name)
    except Exception as e:
        logger.error(f"解析输出路径失败: {e}")
        return 1

    logger.info(f"输入文件: {excel_path}")
    logger.info(f"输出文件: {output_path}")
    logger.info(f"处理子表: {sheet_tasks}")
    logger.info("")

    if args.no_include_pre:
        logger.info("注意: GMC表格将跳过免前(第3行)")
        logger.info("")

    result = process_workbook(excel_path, sheet_tasks, str(output_path), args.verbose)

    return 0 if result else 1


if __name__ == '__main__':
    sys.exit(main())
