# -*- coding: utf-8 -*-
"""
使用 pandas + XlsxWriter 生成 ADR 组合图，避免 openpyxl 绘图导致
Microsoft Excel 报错“Removed Part: /xl/drawings/drawing1.xml”。

流程：用 openpyxl 仅读取源表数据 → 用 pd.ExcelWriter(engine='xlsxwriter') 写入
新工作簿（数据 + 图表），保证交付文件通过 MS Excel 校验，无修复弹窗。
"""

from __future__ import annotations

import argparse
import math
import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
DEFAULT_INPUT = BASE / "input" / "不同剂量组ADR分析 (TFL).xlsx"
DEFAULT_OUTPUT = BASE / "output" / "不同剂量组ADR分析 (TFL).xlsx"

# 供模块被导入时使用的默认输入/输出
TFL = DEFAULT_INPUT
OUT = DEFAULT_OUTPUT

GROUP_NAMES = [
    "低剂量试验组",
    "高剂量试验组",
    "低剂量佐剂组",
    "高剂量佐剂组",
    "安慰剂组",
]


def _adr_total_rows(ws):
    """返回 (adr_rows, total_rows) 列表。"""
    adr_rows, total_rows = [], []
    r = 3
    while r <= ws.max_row:
        if ws.cell(r, 1).value:
            adr_rows.append(r)
        if ws.cell(r + 3, 2).value == "Total":
            total_rows.append(r + 3)
        r += 4
    return adr_rows, total_rows


def _axis_bounds_from_counts(counts_series: list) -> tuple[float, float, float]:
    """根据例数数据计算主 Y 轴（例数）的 min, max, major_unit。"""
    flat = [v for g in counts_series for v in g if v is not None]
    try:
        vals = [float(x) for x in flat]
    except (TypeError, ValueError):
        vals = [0]
    if not vals:
        return 0.0, 10.0, 2.0
    vmin, vmax = min(vals), max(vals)
    vmin = min(0, vmin)
    if vmax <= 0:
        vmax = 10.0
        unit = 2.0
    else:
        # 向上取整到较“整”的数，留约 5% 余量
        pad = max(vmax * 0.05, 1)
        vmax = math.ceil((vmax + pad) / 5) * 5 if vmax + pad > 10 else math.ceil(vmax + pad)
        unit = max(1.0, (vmax - vmin) / 5)
        unit = math.ceil(unit) if unit >= 1 else round(unit * 2) / 2
    return float(vmin), float(vmax), float(unit)


def _axis_bounds_from_rates(rates_series: list) -> tuple[float, float, float]:
    """根据发生率数据（0~1 小数）计算次 Y 轴（发生率%）的 min, max, major_unit。"""
    flat = [v for g in rates_series for v in g if v is not None]
    try:
        vals = [float(x) for x in flat]
    except (TypeError, ValueError):
        vals = [0]
    if not vals:
        return 0.0, 1.0, 0.1
    vmax = max(vals)
    vmin = min(0, min(vals))
    vmax = min(1.0, max(1.0, vmax))
    if vmax <= 0:
        vmax = 1.0
        unit = 0.1
    else:
        pad = 0.02
        vmax = min(1.0, math.ceil((vmax + pad) * 10) / 10.0)
        unit = max(0.05, (vmax - vmin) / 5)
        unit = round(unit * 20) / 20
    return float(vmin), float(vmax), float(unit)


def _axis_bounds_from_duration(values_series: list) -> tuple[float, float, float]:
    """根据持续时间（天）数据计算 Y 轴的 min, max, major_unit。"""
    flat = [v for g in values_series for v in g if v is not None]
    try:
        vals = [float(x) for x in flat]
    except (TypeError, ValueError):
        vals = [0]
    if not vals:
        return 0.0, 10.0, 2.0
    vmin, vmax = min(vals), max(vals)
    vmin = min(0, vmin)
    if vmax <= vmin:
        vmax = vmin + 10
        unit = 2.0
    else:
        pad = max((vmax - vmin) * 0.05, 0.5)
        vmax = math.ceil(vmax + pad)
        unit = max(0.5, (vmax - vmin) / 5)
        unit = math.ceil(unit * 2) / 2
    return float(vmin), float(vmax), float(unit)


# -------------------- 临床配色逻辑（可选启用） --------------------

COLOR_MAP = {
    # 试验组 (蓝色系)
    "低剂量试验组": "#5B9BD5",  # Muted Medium Blue - Standard
    "高剂量试验组": "#254061",  # Deep Navy Blue - Intensity
    # 佐剂/对照组 (红/橙)
    "低剂量佐剂组": "#ED7D31",  # Soft Orange
    "高剂量佐剂组": "#C00000",  # Deep Brick Red
    # 安慰剂组 (中性灰)
    "安慰剂组": "#7F7F7F",      # Medium Gray - Baseline
}

# 非临床配色时的固定色板（与 GROUP_NAMES 顺序一致）
LEGACY_COLORS = [
    "#5B9BD5",  # 低剂量试验组
    "#255E91",  # 高剂量试验组
    "#ED7D31",  # 低剂量佐剂组
    "#C00000",  # 高剂量佐剂组
    "#7F7F7F",  # 安慰剂组
]


def _get_series_color(name: str, use_clinical: bool) -> str:
    """
    根据系列名称返回 HEX 颜色。

    - 当 use_clinical=False 时，保持旧的固定色板（兼容历史）；
    - 当 use_clinical=True 时，使用 CSR 临床配色方案（COLOR_MAP + 名称匹配）。
    """
    if not use_clinical:
        # 按 GROUP_NAMES 顺序回退
        try:
            base = name.split("-", 1)[0]
            idx = GROUP_NAMES.index(base)
            return LEGACY_COLORS[idx]
        except Exception:
            return "#7F7F7F"

    # 临床配色逻辑
    if not name:
        return COLOR_MAP["安慰剂组"]

    raw = name.strip()
    base = raw.split("-", 1)[0]

    if base in COLOR_MAP:
        return COLOR_MAP[base]

    n = raw
    if "试验" in n and "低剂量" in n:
        return COLOR_MAP["低剂量试验组"]
    if "试验" in n and "高剂量" in n:
        return COLOR_MAP["高剂量试验组"]

    if ("佐剂" in n or "对照" in n) and "低剂量" in n:
        return COLOR_MAP["低剂量佐剂组"]
    if ("佐剂" in n or "对照" in n) and "高剂量" in n:
        return COLOR_MAP["高剂量佐剂组"]

    if "安慰剂" in n or "placebo" in n.lower():
        return COLOR_MAP["安慰剂组"]

    return COLOR_MAP["安慰剂组"]


def _read_source_data():
    """
    用 openpyxl 仅读取数据，不写入任何图表。
    
    Returns:
        (sheet_name, rows, chart_data, max_col) 元组
        
    Raises:
        FileNotFoundError: 源文件不存在
        ValueError: 文件格式错误或无法读取
    """
    if not TFL.exists():
        raise FileNotFoundError(f"源 Excel 文件不存在: {TFL}")
    
    wb = None
    wb_data = None
    try:
        wb = load_workbook(TFL, read_only=False, data_only=False)
        ws = wb.active
        sheet_name = ws.title

        # 整表：保留公式或值（用于原样写回）
        max_row, max_col = ws.max_row, ws.max_column
        rows = []
        for r in range(1, max_row + 1):
            row = []
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                row.append(cell.value)
            rows.append(row)

        # 图表用数值：
        # - ADR发生情况（原版）：按 4 行一组 + Total 行提取 counts/rates
        # - ADR持续时间（本次输入）：表格结构为 A列ADR，B-F为5组的“天数/中位数”等数值
        wb_data = load_workbook(TFL, read_only=False, data_only=True)
        ws_data = wb_data.active

        adr_rows, total_rows = _adr_total_rows(ws_data)
        if total_rows:
            chart_data = {
                "kind": "incidence",
                "adr_names": [],
                "counts": [[] for _ in range(5)],
                "rates": [[] for _ in range(5)],
            }
            for ar, tr in zip(adr_rows, total_rows):
                chart_data["adr_names"].append(ws_data.cell(ar, 1).value or "")
                for i, col in enumerate([3, 6, 9, 12, 15]):
                    v = ws_data.cell(tr, col).value
                    chart_data["counts"][i].append(v if v is not None else 0)
                for i, col in enumerate([5, 8, 11, 14, 17]):
                    v = ws_data.cell(tr, col).value
                    if v is None:
                        v = 0
                    elif isinstance(v, (int, float)) and v > 1:
                        v = v / 100.0
                    chart_data["rates"][i].append(v)
        else:
            # Duration table: header row has group names in B-F, row2 has units, data from row3
            group_names = []
            for c in range(2, 7):  # B..F
                v = ws_data.cell(1, c).value
                group_names.append(str(v).strip() if v is not None else f"Group{c-1}")
            adr_names = []
            values = [[] for _ in range(5)]
            for r in range(3, ws_data.max_row + 1):
                adr = ws_data.cell(r, 1).value
                if adr is None or str(adr).strip() == "":
                    continue
                adr_names.append(str(adr).strip())
                for i, c in enumerate(range(2, 7)):
                    v = ws_data.cell(r, c).value
                    try:
                        values[i].append(float(v) if v is not None else 0.0)
                    except Exception:
                        values[i].append(0.0)
            chart_data = {
                "kind": "duration",
                "adr_names": adr_names,
                "group_names": group_names,
                "values": values,
            }
        
        return sheet_name, rows, chart_data, max_col
        
    except Exception as e:
        raise ValueError(f"读取源文件失败: {e}") from e
    finally:
        if wb_data is not None:
            wb_data.close()
        if wb is not None:
            wb.close()


def build(use_clinical_colors: bool = False):
    """
    构建 Excel 图表。
    
    Raises:
        FileNotFoundError: 源文件不存在
        ValueError: 数据格式错误
        OSError: 文件操作失败
    """
    try:
        sheet_name, rows, chart_data, max_col = _read_source_data()
    except FileNotFoundError:
        raise
    except Exception as e:
        raise ValueError(f"读取源数据失败: {e}") from e
    
    n_cat = len(chart_data.get("adr_names", []))
    if n_cat == 0:
        raise ValueError("未找到任何 ADR 数据")

    # 避免文件占用导致写入不完整
    if OUT.exists():
        try:
            os.remove(OUT)
        except OSError as e:
            print(f"警告: 无法删除旧文件 {OUT}: {e}")
            # 继续尝试写入，可能会失败但至少尝试一下

    with pd.ExcelWriter(str(OUT), engine="xlsxwriter", engine_kwargs={"options": {"nan_inf_to_errors": True}}) as writer:
        workbook = writer.book
        pct_fmt = workbook.add_format({"num_format": "0.00%"})
        # 主表：原样写回数据（公式或值）
        ws = workbook.add_worksheet(sheet_name)
        for r_idx, row in enumerate(rows):
            for c_idx, val in enumerate(row):
                if val is None:
                    continue
                if isinstance(val, str) and val.startswith("="):
                    ws.write_formula(r_idx, c_idx, val, pct_fmt if c_idx in [4, 7, 10, 13, 16] else None)
                else:
                    ws.write(r_idx, c_idx, val, pct_fmt if c_idx in [4, 7, 10, 13, 16] else None)

        kind = chart_data.get("kind", "incidence")

        # ChartData 表（隐藏需 Excel 支持，此处不隐藏以保证兼容性）
        cd = workbook.add_worksheet("ChartData")
        cd.write(0, 0, "ADR")

        if kind == "incidence":
            for i, name in enumerate(GROUP_NAMES):
                cd.write(0, 1 + i, f"{name}-例数")
            for i, name in enumerate(GROUP_NAMES):
                cd.write(0, 6 + i, f"{name}-发生率")
            for j in range(n_cat):
                cd.write(1 + j, 0, chart_data["adr_names"][j])
                for i in range(5):
                    cd.write(1 + j, 1 + i, chart_data["counts"][i][j])
                for i in range(5):
                    cd.write(1 + j, 6 + i, chart_data["rates"][i][j])

            # 组合图：主图 column，次图 line 用 combine 合并，折线系列指定 y2_axis
            col_chart = workbook.add_chart({"type": "column"})
            line_chart = workbook.add_chart({"type": "line"})

            # 柱状图 5 个系列（例数）
            for i in range(5):
                col_letter = chr(66 + i)  # B,C,D,E,F
                series_name = f"{GROUP_NAMES[i]}-例数"
                color = _get_series_color(series_name, use_clinical_colors)
                col_chart.add_series({
                    "name": f"=ChartData!${col_letter}$1",
                    "categories": f"=ChartData!$A$2:$A${1 + n_cat}",
                    "values": f"=ChartData!${col_letter}$2:${col_letter}${1 + n_cat}",
                    "fill": {"color": color},
                    "border": {"color": "#FFFFFF", "width": 0.5} if use_clinical_colors else {"color": "black"},
                })

            # 折线 5 个系列（发生率，次坐标 y2_axis）
            for i in range(5):
                col_letter = chr(71 + i)  # G,H,I,J,K
                series_name = f"{GROUP_NAMES[i]}-发生率"
                color = _get_series_color(series_name, use_clinical_colors)
                line_kwargs = {
                    "name": f"=ChartData!${col_letter}$1",
                    "categories": f"=ChartData!$A$2:$A${1 + n_cat}",
                    "values": f"=ChartData!${col_letter}$2:${col_letter}${1 + n_cat}",
                    "y2_axis": 1,
                }
                if use_clinical_colors:
                    line_kwargs.update({
                        "line": {"width": 1.5, "color": color},
                        "marker": {
                            "type": "circle",
                            "size": 5,
                            "border": {"color": color},
                            "fill": {"color": color},
                        },
                    })
                else:
                    line_kwargs.update({
                        "line": {"width": 1.25, "color": LEGACY_COLORS[i]},
                        "marker": {"type": "circle", "size": 5},
                    })
                line_chart.add_series(line_kwargs)

            col_chart.combine(line_chart)
            chart = col_chart

            # 坐标轴范围与表格数据匹配
            y_min, y_max, y_unit = _axis_bounds_from_counts(chart_data["counts"])
            r_min, r_max, r_unit = _axis_bounds_from_rates(chart_data["rates"])

            title = "不同剂量组ADR发生情况（临床规范配色）" if use_clinical_colors else "不同剂量组ADR发生情况"
            chart.set_title({"name": title})
            chart.set_x_axis({
                "name": "",
                "major_tick_mark": "outside",
                "line": {"color": "black", "width": 0.75},
            })
            chart.set_y_axis({
                "name": "例数 (n)",
                "major_gridlines": {"visible": False},
                "major_tick_mark": "outside",
                "line": {"color": "black", "width": 0.75},
                "num_format": "0",
                "name_font": {"size": 10, "name": "Arial"},
                "min": y_min,
                "max": y_max,
                "major_unit": y_unit,
            })
            line_chart.set_y2_axis({
                "name": "发生率 (%)",
                "major_gridlines": {"visible": False},
                "major_tick_mark": "outside",
                "line": {"color": "black", "width": 0.75},
                "num_format": "0.00%",
                "min": r_min,
                "max": r_max,
                "major_unit": r_unit,
                "name_font": {"size": 10, "name": "Arial"},
            })
            chart.set_legend({"position": "bottom"})
            ws.insert_chart("L6", chart)
        else:
            # duration chart: clustered columns (days)
            group_names = chart_data.get("group_names", GROUP_NAMES)
            for i, name in enumerate(group_names[:5]):
                cd.write(0, 1 + i, str(name))
            for j in range(n_cat):
                cd.write(1 + j, 0, chart_data["adr_names"][j])
                for i in range(5):
                    cd.write(1 + j, 1 + i, chart_data["values"][i][j])

            chart = workbook.add_chart({"type": "column"})
            for i in range(5):
                col_letter = chr(66 + i)  # B..F
                chart.add_series({
                    "name": f"=ChartData!${col_letter}$1",
                    "categories": f"=ChartData!$A$2:$A${1 + n_cat}",
                    "values": f"=ChartData!${col_letter}$2:${col_letter}${1 + n_cat}",
                    "fill": {"color": LEGACY_COLORS[i]},
                    "border": {"color": "black"},
                })

            # 坐标轴范围与表格数据匹配
            dy_min, dy_max, dy_unit = _axis_bounds_from_duration(chart_data["values"])

            chart.set_title({"name": "不同剂量组ADR持续时间"})
            chart.set_x_axis({"line": {"color": "black", "width": 0.75}})
            chart.set_y_axis({
                "name": "持续时间（天）",
                "major_gridlines": {"visible": False},
                "line": {"color": "black", "width": 0.75},
                "num_format": "0.0",
                "name_font": {"size": 10, "name": "Arial"},
                "min": dy_min,
                "max": dy_max,
                "major_unit": dy_unit,
            })
            chart.set_legend({"position": "bottom"})
            ws.insert_chart("H3", chart)

    print("charts built with XlsxWriter:", OUT)


if __name__ == "__main__":
    try:
        parser = argparse.ArgumentParser(description="使用 XlsxWriter 在 Excel 中生成 ADR 组合图")
        parser.add_argument(
            "--input",
            "-i",
            default=str(DEFAULT_INPUT),
            help="输入 Excel 路径（默认：01_Excel_Charts/input/不同剂量组ADR分析 (TFL).xlsx）",
        )
        parser.add_argument(
            "--output",
            "-o",
            default=str(DEFAULT_OUTPUT),
            help="输出 Excel 路径（默认：01_Excel_Charts/output/不同剂量组ADR分析 (TFL).xlsx）",
        )
        parser.add_argument(
            "--clinical-colors",
            action="store_true",
            help="启用临床发表/CSR 规范配色（试验组蓝系、佐剂组红橙、安慰剂组灰色）",
        )
        args = parser.parse_args()

        # 运行时设置输入输出（供 _read_source_data / build 读取）
        TFL = Path(args.input)
        OUT = Path(args.output)
        OUT.parent.mkdir(parents=True, exist_ok=True)

        build(use_clinical_colors=args.clinical_colors)
    except FileNotFoundError as e:
        print(f"错误: 文件未找到 - {e}")
        exit(1)
    except ValueError as e:
        print(f"错误: 数据格式错误 - {e}")
        exit(1)
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()
        exit(1)
