"""
Excel 子表 <-> Word 表格映射：数据结构、文本匹配、plan 生成与 JSON 加载。

主程序 `word_to_excel_to_figure.py` 从本包导入上述能力，避免双轨实现。本文件位于 `lib/`，勿直接运行。
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


@dataclass(frozen=True)
class CellRange:
    sheet: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int

    @property
    def nrows(self) -> int:
        return self.max_row - self.min_row + 1

    @property
    def ncols(self) -> int:
        return self.max_col - self.min_col + 1


@dataclass(frozen=True)
class ChartSeriesRefs:
    cat: CellRange
    val: CellRange


@dataclass(frozen=True)
class WordTableRef:
    """指定 Word 文档中的某一个表（doc.Content.Tables 的 1-based 索引）。"""

    word_path: str
    table_index: int


def strip_word_cell_text(text: Any) -> str:
    if text is None:
        return ""
    s = str(text)
    s = s.replace("\x07", "").replace("\r", "").replace("\x0b", "")
    s = s.strip()
    s = re.sub(r"\s+", "", s)
    if s.startswith("'"):
        s = s.lstrip("'")
    return s


def label_tokens(s: str) -> list[str]:
    s_norm = strip_word_cell_text(s).upper()
    return re.findall(r"[A-Z0-9]+", s_norm)


def labels_match(cell_text: Any, target_text: Any) -> bool:
    tgt = strip_word_cell_text(target_text)
    if not tgt:
        return False
    cell = strip_word_cell_text(cell_text)
    if cell == tgt:
        return True
    if tgt in cell or cell in tgt:
        return True
    cell_up = cell.upper()
    toks = label_tokens(tgt)
    return bool(toks) and all(t in cell_up for t in toks)


def subtable_id_from_cell_range(r: CellRange) -> str:
    return f"{r.sheet}|{r.min_row}|{r.max_row}|{r.min_col}|{r.max_col}"


def load_table_mapping_json(table_map_json_path: str | None) -> dict[str, list[WordTableRef]]:
    """
    从 JSON 读取你确认后的 table 映射。
    若 path 为 None 或空字符串，返回空 dict。
    """
    if not table_map_json_path:
        return {}

    p = Path(table_map_json_path).expanduser().resolve()
    if not p.is_file():
        raise FileNotFoundError(f"--table-map-json 指定文件不存在: {p}")
    data = json.loads(p.read_text(encoding="utf-8"))
    subtables = data.get("subtables", {}) or {}
    mapping: dict[str, list[WordTableRef]] = {}

    for sid, cand_list in subtables.items():
        if not cand_list:
            continue
        selected = [c for c in cand_list if bool(c.get("selected"))]
        if not selected:
            selected = cand_list[:1]
        mapping[sid] = [
            WordTableRef(
                word_path=str(c.get("word_file")),
                table_index=int(c.get("table_index")),
            )
            for c in selected
        ]

    return mapping


def build_table_mapping_plan(
    template_path: Path,
    word_parts: list[Path],
    series_refs: list[ChartSeriesRefs],
    *,
    top_k_per_subtable: int = 3,
) -> dict[str, Any]:
    """
    生成“Excel 子表 -> Word 具体表”的候选映射（供人工确认）。
    """
    from word_to_excel_to_figure import WordComRunner  # 延迟导入避免循环依赖

    template_wb = openpyxl.load_workbook(str(template_path), data_only=False)
    try:
        subtables: dict[str, list[ChartSeriesRefs]] = {}
        for sr in series_refs:
            sid = subtable_id_from_cell_range(sr.cat)
            subtables.setdefault(sid, []).append(sr)

        sub_labels: dict[str, list[str]] = {}
        sub_cat_samples: dict[str, list[str]] = {}
        for sid, srs in subtables.items():
            sr0 = srs[0]
            ws_cat = template_wb[sr0.cat.sheet]
            cat_values = [ws_cat.cell(sr0.cat.min_row + k, sr0.cat.min_col).value for k in range(sr0.cat.nrows)]
            cat_norm = [strip_word_cell_text(v) for v in cat_values]
            cat_norm = [x for x in cat_norm if x]
            sub_labels[sid] = cat_norm
            sub_cat_samples[sid] = cat_norm[:5]
    finally:
        try:
            template_wb.close()
        except Exception:
            pass

    table_infos: dict[str, list[dict[str, Any]]] = {}
    with WordComRunner() as runner:
        for wp in word_parts:
            doc = runner.open_doc(wp)
            infos: list[dict[str, Any]] = []
            tables = doc.Content.Tables
            for ti in range(1, tables.Count + 1):
                t = tables.Item(ti)
                try:
                    raw = str(t.Range.Text)
                except Exception:
                    raw = ""
                table_text = strip_word_cell_text(raw)
                snippet = raw.replace("\r", " ").replace("\n", " ")
                snippet = re.sub(r"\s+", " ", snippet).strip()
                snippet = snippet[:140] if len(snippet) > 140 else snippet
                infos.append({"table_index": ti, "snippet": snippet, "table_text": table_text})
            table_infos[str(wp)] = infos

    subtables_out: dict[str, list[dict[str, Any]]] = {}
    for sid, cat_labels in sub_labels.items():
        sample_labels = cat_labels[:6]
        candidates: list[tuple[int, str, int, str]] = []
        for wp_str, infos in table_infos.items():
            for info in infos:
                table_text = info.get("table_text", "")
                score = sum(1 for lbl in sample_labels if labels_match(table_text, lbl))
                candidates.append((score, wp_str, int(info["table_index"]), str(info.get("snippet", ""))))

        candidates.sort(key=lambda x: (-x[0], x[2]))
        top = candidates[:top_k_per_subtable]
        subtables_out[sid] = [
            {
                "word_file": wp_str,
                "table_index": ti,
                "score": score,
                "snippet": snippet,
            }
            for score, wp_str, ti, snippet in top
        ]

    return {
        "template_path": str(template_path),
        "subtables": subtables_out,
        "subtable_cat_samples": sub_cat_samples,
        "top_k_per_subtable": top_k_per_subtable,
        "note": "你需要从每个 subtable 的候选列表里选择真正对应的 Word 表格（table_index）。然后把你的选择写成 table-map-json 供程序二次运行。",
    }
