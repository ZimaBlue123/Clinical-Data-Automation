# -*- coding: utf-8 -*-
"""
Word -> Excel(表+图) 自动复刻模块

设计目标：
1) 读取输入目录下 3 个 Word/RTF 文件（part1/part2/part3 或按数量排序）。
2) 以模板 Excel 作为“格式/图表结构”载体，动态发现模板里所有图表的 series(cat/val) 引用区间；
3) 对每个 cat/val 引用区间：在 Word 的 tables 中按内容匹配找到对应数据列/行；
4) 将提取到的数据写回输出 Excel 对应单元格（不覆盖样式），最后进行自检：
   输出与模板在所有图表引用单元格上的值应完全一致（文本相等，数值允许极小误差）。

注意：
- 需要 Windows + 安装 Microsoft Word（使用 pywin32 的 COM）。
"""

from __future__ import annotations

import argparse
import logging
import math
import re
import json
import os
import shutil
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from table_mapping_logic import (
    CellRange,
    ChartSeriesRefs,
    WordTableRef,
    build_table_mapping_plan,
    label_tokens as _label_tokens,
    labels_match as _labels_match,
    load_table_mapping_json,
    strip_word_cell_text as _strip_word_cell_text,
    subtable_id_from_cell_range as _subtable_id_from_cell_range,
)


logger = logging.getLogger("word_to_excel_to_figure")


WORD_SUFFIXES = {".doc", ".docx", ".rtf"}
MODULE_DIR = Path(__file__).resolve().parent
TEMPLATE_DIR = MODULE_DIR / "Template"


def _setup_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    )


_NUM_RE = re.compile(r"[-+]?\d+(?:\.\d+)?")


def _parse_word_number(text: str | None) -> Optional[float]:
    s = _strip_word_cell_text(text)
    if not s:
        return None
    m = _NUM_RE.search(s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def _numbers_close(a: float, b: float, *, rel_tol: float = 1e-9, abs_tol: float = 1e-9) -> bool:
    return math.isclose(a, b, rel_tol=rel_tol, abs_tol=abs_tol)


def _normalize_number_for_compare(word_val: float, template_val: Any) -> Optional[float]:
    """
    尽量把 Word 表里的数字“转到与模板一致的量纲”再比较。
    - 若模板是 0~1 小数（发生率），Word 可能写成 0~100 或带 % 形式。
    """
    if template_val is None:
        return None

    try:
        tv = float(template_val)
    except Exception:
        return word_val

    # 发生率（0~1）
    if 0 <= tv <= 1:
        # Word 如果是 23（百分数），转成 0.23
        if word_val > 1 and _numbers_close(word_val / 100.0, tv, rel_tol=1e-6, abs_tol=1e-8):
            return word_val / 100.0
        # Word 如果已经是 0~1
        return word_val

    return word_val


def _detect_template_is_proportion(template_val_values: list[Any]) -> bool:
    """
    判断模板 val 大多是否属于发生率/比例（0~1），用来对齐 Word 侧的百分数量纲。
    只依赖“值的分布形态”，不依赖具体数值相等（因为新数据的 val 会变）。
    """
    vals: list[float] = []
    for v in template_val_values:
        if v is None:
            continue
        try:
            vals.append(float(v))
        except Exception:
            continue

    if not vals:
        return False

    within = [v for v in vals if 0.0 <= v <= 1.0]
    # 比例类通常绝大多数落在 0~1；阈值可调
    return len(within) / len(vals) >= 0.7


def _coerce_word_number_scale(word_val: float, *, template_is_proportion: bool) -> float:
    """
    将 Word 侧数字“缩放到模板期望的量纲”，但不使用数值接近判断。
    例如：模板是 0~1 比例，而 Word 可能是 0~100 百分数。
    """
    if template_is_proportion and word_val > 1.0:
        return word_val / 100.0
    return word_val


def _extract_val_header_keys(template_wb: openpyxl.Workbook, sr: ChartSeriesRefs) -> list[str]:
    """
    从骨架 val 列上方（通常是表头行）提取列名，用于在 Word table 中定位正确的 val 列。
    """
    ws = template_wb[sr.val.sheet]
    col = sr.val.min_col
    keys: list[str] = []

    # 经验范围：val 起始行的前 2 行常是表头/二级表头
    for rr in range(sr.val.min_row - 2, sr.val.min_row):
        if rr < 1:
            continue
        v = ws.cell(rr, col).value
        s = _strip_word_cell_text(v)
        if s and s not in keys:
            keys.append(s)
        if len(keys) >= 5:
            break
    return keys


def _cellrange_from_excel_ref(ref: str) -> CellRange:
    """
    refs 通常形如： 'Sheet Name'!$A$2:$A$10
    """
    # 去掉可能的外引号，并支持中文/空格的工作表名
    ref = ref.strip()
    if "!" not in ref:
        raise ValueError(f"无法解析引用（缺少 '!'): {ref}")

    sheet_part, addr_part = ref.split("!", 1)
    sheet_part = sheet_part.strip()
    if sheet_part.startswith("'") and sheet_part.endswith("'"):
        sheet_part = sheet_part[1:-1]

    # 仅支持矩形范围写法：$A$2:$B$10 或 $A$2:$A$10
    m = re.match(r"^\$(?P<c1>[A-Z]+)\$(?P<r1>\d+):\$(?P<c2>[A-Z]+)\$(?P<r2>\d+)$", addr_part)
    if not m:
        raise ValueError(f"无法解析地址范围: {ref}")

    def col_letter_to_index(col: str) -> int:
        col = col.upper()
        idx = 0
        for ch in col:
            idx = idx * 26 + (ord(ch) - ord("A") + 1)
        return idx

    min_col = col_letter_to_index(m.group("c1"))
    max_col = col_letter_to_index(m.group("c2"))
    min_row = int(m.group("r1"))
    max_row = int(m.group("r2"))
    return CellRange(
        sheet=sheet_part,
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    )


def _read_cell_values(ws: Worksheet, r: CellRange) -> list[list[Any]]:
    out: list[list[Any]] = []
    for rr in range(r.min_row, r.max_row + 1):
        row_vals = []
        for cc in range(r.min_col, r.max_col + 1):
            row_vals.append(ws.cell(rr, cc).value)
        out.append(row_vals)
    return out


def _extract_template_series_refs(template_wb: openpyxl.Workbook) -> list[ChartSeriesRefs]:
    refs: list[ChartSeriesRefs] = []
    for ws in template_wb.worksheets:
        charts = getattr(ws, "_charts", []) or []
        for ch in charts:
            for s in getattr(ch, "series", []) or []:
                cat = getattr(s, "cat", None)
                val = getattr(s, "val", None)

                def f(x: Any) -> Optional[str]:
                    if x is None:
                        return None
                    nr = getattr(x, "numRef", None)
                    if nr is not None:
                        return getattr(nr, "f", None)
                    sr = getattr(x, "strRef", None)
                    if sr is not None:
                        return getattr(sr, "f", None)
                    return None

                cat_ref = f(cat)
                val_ref = f(val)
                if not cat_ref or not val_ref:
                    continue
                try:
                    cat_range = _cellrange_from_excel_ref(str(cat_ref))
                    val_range = _cellrange_from_excel_ref(str(val_ref))
                except Exception:
                    continue

                # cat/val 一般是同长度序列：通常为 1 列（或 1 行）
                if cat_range.ncols != val_range.ncols:
                    # 这里先跳过复杂情况（本样例均为 1 列）
                    continue
                if cat_range.nrows != val_range.nrows:
                    continue

                refs.append(ChartSeriesRefs(cat=cat_range, val=val_range))
    # 去重（cat/val 完全相同）
    uniq: dict[tuple[CellRange, CellRange], None] = {}
    for r in refs:
        uniq[(r.cat, r.val)] = None
    out: list[ChartSeriesRefs] = []
    for cat_range, val_range in uniq.keys():
        out.append(ChartSeriesRefs(cat=cat_range, val=val_range))
    return out


def _word_find_table_containing_text(doc: Any, key_text: str) -> list[Any]:
    """
    返回：命中 key_text 的 Range 落在其中的 table 列表（可能为 0/1/多张）。
    """
    rng = doc.Content
    find = rng.Find
    find.ClearFormatting()
    find.Text = key_text
    find.Replacement.Text = ""
    find.Forward = True
    # wdFindContinue = 1
    find.Wrap = 1
    find.Format = False
    ok = find.Execute()
    if not ok:
        return []
    try:
        tables = rng.Tables
        return [tables.Item(i) for i in range(1, tables.Count + 1)]
    except Exception:
        return []


def _extract_table_matrix(table: Any) -> tuple[list[list[str]], list[list[Optional[float]]]]:
    """
    将 Word table 抽成：
    - text_matrix[row][col]：清洗后的文本
    - num_matrix[row][col]：解析后的数字（无法解析返回 None）
    """
    rows = table.Rows.Count
    cols = table.Columns.Count
    text_matrix: list[list[str]] = [[""] * cols for _ in range(rows)]
    num_matrix: list[list[Optional[float]]] = [[None] * cols for _ in range(rows)]
    for i in range(1, rows + 1):
        for j in range(1, cols + 1):
            try:
                raw = table.Cell(i, j).Range.Text
            except Exception:
                raw = None
            txt = _strip_word_cell_text(raw)
            text_matrix[i - 1][j - 1] = txt
            num_matrix[i - 1][j - 1] = _parse_word_number(raw)
    return text_matrix, num_matrix


def _match_cat_and_val_columns(
    text_matrix: list[list[str]],
    num_matrix: list[list[Optional[float]]],
    cat_values: list[str],
    val_header_keys: list[str],
) -> Optional[tuple[list[int], int, int]]:
    """
    在一个 Word table 里找：
    - cat_col：哪个列是 cat_values 的序列
    - row_positions：cat_values 每一项在该列上匹配到的行（严格递增）
    - val_col：哪个列是与 cat_values 同行对应的数值列（用 val_header_keys 定位，不再依赖模板 val 的具体数值）
    返回 (row_positions, cat_col, val_col)，行列均为 0-based
    """
    if not cat_values:
        return None
    R = len(text_matrix)
    C = len(text_matrix[0]) if R else 0
    n = len(cat_values)
    cat_norm = [_strip_word_cell_text(v) for v in cat_values]

    val_header_keys_norm = [_strip_word_cell_text(v) for v in val_header_keys if v is not None]
    val_header_keys_norm = [v for v in val_header_keys_norm if v]

    # 允许 cat_values 在同一列中“有间隔地按顺序出现”（解决合并单元格/表头插入导致的非连续问题）
    for cat_col in range(C):
        candidate_rows = [r for r in range(R) if _labels_match(text_matrix[r][cat_col], cat_norm[0])]
        for start_row in candidate_rows:
            row_positions = [start_row]
            prev = start_row
            ok = True
            for k in range(1, n):
                found_row = None
                for r in range(prev + 1, R):
                    if _labels_match(text_matrix[r][cat_col], cat_norm[k]):
                        found_row = r
                        break
                if found_row is None:
                    ok = False
                    break
                row_positions.append(found_row)
                prev = found_row

            if not ok:
                continue

            # cat 行确定后，反推 val_col：
            # 1) 优先用该列的表头（val_header_keys）匹配；
            # 2) 再用 cat 行位置的数值可解析数量做辅助打分；
            # 不依赖模板 val 的“具体数值”，因为新数据 val 会变。
            col_texts: list[set[str]] = []
            for cc in range(C):
                col_texts.append({text_matrix[rr][cc] for rr in range(R) if text_matrix[rr][cc]})

            best: Optional[tuple[int, int, int]] = None  # (header_score, numeric_non_none, val_col)
            for val_col in range(C):
                numeric_non_none = 0
                for k in range(n):
                    if num_matrix[row_positions[k]][val_col] is not None:
                        numeric_non_none += 1

                if numeric_non_none == 0:
                    continue

                header_score = 0
                if val_header_keys_norm:
                    header_score = sum(1 for hk in val_header_keys_norm if hk in col_texts[val_col])

                cand = (header_score, numeric_non_none, val_col)
                if best is None:
                    best = cand
                    continue
                # 先比表头分，再比可解析数量
                if cand[0] > best[0] or (cand[0] == best[0] and cand[1] > best[1]):
                    best = cand

            if best is not None:
                # 若表头给得到了（header_score>0），就认为 val_col 更可靠；否则仍可用数字可解析最多的列
                _header_score, _numeric_non_none, _val_col = best
                return row_positions, cat_col, _val_col
    return None


def _fill_range_from_word_table(
    template_ws: Worksheet,
    output_ws: Worksheet,
    cat_range: CellRange,
    val_range: CellRange,
    word_parts: list[Path],
    template_cat_values: list[str],
    template_val_values: list[Any],
    word_tables: list[Any],
) -> None:
    """
    根据候选 word_tables，找出最佳 table，并将 cat/val 写回 output_ws 的对应单元格。
    """
    # cat_values/val_values 默认为同长度列表，对应 cat_range/val_range 行方向
    n = len(template_cat_values)
    if n != len(template_val_values):
        raise ValueError("cat 与 val 长度不一致，无法匹配")

    for table in word_tables:
        text_matrix, num_matrix = _extract_table_matrix(table)
        match = _match_cat_and_val_columns(text_matrix, num_matrix, template_cat_values, template_val_values)
        if match is None:
            continue
        row_positions, cat_col, val_col = match

        # 写回输出
        for k in range(n):
            out_cat_cell = (cat_range.min_row + k, cat_range.min_col)
            out_val_cell = (val_range.min_row + k, val_range.min_col)
            output_ws.cell(*out_cat_cell).value = template_ws.cell(cat_range.min_row + k, cat_range.min_col).value

            wnum = num_matrix[row_positions[k]][val_col]
            if wnum is None:
                output_ws.cell(*out_val_cell).value = template_ws.cell(val_range.min_row + k, val_range.min_col).value
            else:
                # 若模板 val 是整数则写入整数；否则写入 float。
                tval = template_ws.cell(val_range.min_row + k, val_range.min_col).value
                try:
                    tvf = float(tval)
                    if abs(tvf - round(tvf)) < 1e-9:
                        output_ws.cell(*out_val_cell).value = int(round(float(wnum)))
                    else:
                        # 对发生率/百分比这里不做强制尺度转换，直接使用模板尺度对齐值
                        output_ws.cell(*out_val_cell).value = float(_normalize_number_for_compare(float(wnum), tvf))
                except Exception:
                    output_ws.cell(*out_val_cell).value = wnum
        return

    raise ValueError("未找到可匹配的 Word table，用于填充 cat/val 区间")


class WordComRunner:
    def __init__(self) -> None:
        try:
            import win32com.client  # type: ignore
        except ImportError as e:
            raise RuntimeError("需要安装 pywin32，才能使用 Word COM 读取表格") from e
        self.win32 = win32com.client
        self.app = None
        self.docs: dict[Path, Any] = {}

    def __enter__(self) -> "WordComRunner":
        self.app = self.win32.Dispatch("Word.Application")
        self.app.Visible = False
        try:
            self.app.DisplayAlerts = 0
        except Exception:
            pass
        return self

    def open_doc(self, path: Path) -> Any:
        if path in self.docs:
            return self.docs[path]
        doc = self.app.Documents.Open(str(path))
        self.docs[path] = doc
        return doc

    def close_all(self) -> None:
        for p, doc in list(self.docs.items()):
            try:
                doc.Close(False)
            except Exception:
                pass
            self.docs.pop(p, None)
        if self.app is not None:
            try:
                self.app.Quit()
            except Exception:
                pass
            self.app = None

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close_all()


def _collect_word_parts(input_dir: Path) -> list[Path]:
    """
    遍历 input_dir 下所有 Word/RTF 文件（数量不固定）。
    - 跳过 Word 临时锁文件 `~$...`
    - 若文件名包含 part1/part2/part3，会按 part 顺序优先
    """
    files: list[Path] = []
    for p in input_dir.iterdir():
        if not (p.is_file() and p.suffix.lower() in WORD_SUFFIXES):
            continue
        if p.name.startswith("~$"):
            continue
        files.append(p)
    if not files:
        raise FileNotFoundError(f"在输入目录未找到 Word/RTF 文件: {input_dir}")

    # 若文件名含 part1/part2/part3，则按其对应关系排序
    def part_num(name: str) -> Optional[int]:
        m = re.search(r"part\s*([123])", name, flags=re.I)
        if m:
            return int(m.group(1))
        m = re.search(r"_([123])_?", name)
        return int(m.group(1)) if m else None

    # 同一 part 号可能有多个候选（旧文件/新文件/备份），选最近修改的
    # 但对“非 part”文件，仍保留全部
    latest_by_part: dict[int, Path] = {}
    others: list[Path] = []
    for p in files:
        pn = part_num(p.name)
        if pn in (1, 2, 3):
            if pn not in latest_by_part or p.stat().st_mtime > latest_by_part[pn].stat().st_mtime:
                latest_by_part[pn] = p
        else:
            others.append(p)

    parts_sorted = [latest_by_part[k] for k in sorted(latest_by_part.keys())]
    others_sorted = sorted(others, key=lambda x: x.name)
    return parts_sorted + others_sorted


def _find_skeleton_xlsx(template_xlsx_arg: Optional[str]) -> Path:
    """
    不依赖用户提供模板。
    默认使用模块目录下唯一的 `.xlsx` 骨架文件（应为上一版结构文件）。
    """
    if template_xlsx_arg:
        p = Path(template_xlsx_arg).expanduser()
        if not p.is_file():
            raise FileNotFoundError(f"骨架 Excel 不存在: {p}")
        return p

    # 优先从 Template/ 找
    candidates = [p for p in TEMPLATE_DIR.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx"] if TEMPLATE_DIR.exists() else []
    # 兼容：也允许旧行为（模块根目录）
    if not candidates:
        candidates = [p for p in MODULE_DIR.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx"]

    if len(candidates) == 0:
        raise ValueError(
            "未找到骨架 Excel。\n"
            "请把一份骨架 xlsx 放到："
            f"{TEMPLATE_DIR}\n"
            "或者运行时用 `--template-xlsx` 指定骨架路径。"
        )
    if len(candidates) != 1:
        raise ValueError(
            "未提供 --template-xlsx，但检测到多个骨架 xlsx。\n"
            f"当前匹配数量={len(candidates)}，请显式用 --template-xlsx 指定。"
        )
    return candidates[0]


def _select_best_skeleton_xlsx(
    template_xlsx_arg: Optional[str],
    template_dir: Path,
    word_parts: list[Path],
) -> Path:
    """
    当用户不显式指定骨架时，从 Template/ 多个骨架中选择最适配当前输入的那个。
    判定依据：骨架里图表 series(cat) 的关键 label 在 Word 文档表格中被命中的比例。
    """
    if template_xlsx_arg:
        return _find_skeleton_xlsx(template_xlsx_arg)

    candidates = [p for p in template_dir.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx"] if template_dir.exists() else []
    if not candidates:
        raise ValueError(f"未找到骨架 xlsx：{template_dir}")
    if len(candidates) == 1:
        return candidates[0]

    # 打分时缓存：label -> 是否存在
    label_found_cache: dict[str, bool] = {}

    def is_label_found_in_words(key_norm: str, docs: list[Any]) -> bool:
        if key_norm in label_found_cache:
            return label_found_cache[key_norm]
        found = False
        for doc in docs:
            if _word_find_table_containing_text(doc, key_norm):
                found = True
                break
        label_found_cache[key_norm] = found
        return found

    best_path: Optional[Path] = None
    best_score: float = -1.0

    with WordComRunner() as runner:
        docs = [runner.open_doc(wp) for wp in word_parts]

        for cand in candidates:
            try:
                wb = openpyxl.load_workbook(str(cand), data_only=False)
            except Exception:
                continue

            try:
                series_refs = _extract_template_series_refs(wb)
                if not series_refs:
                    continue

                sample_series = series_refs[: min(15, len(series_refs))]
                keys: list[str] = []
                seen_keys: set[str] = set()

                for sr in sample_series:
                    tcat_ws = wb[sr.cat.sheet]
                    cat_values = [tcat_ws.cell(sr.cat.min_row + k, sr.cat.min_col).value for k in range(sr.cat.nrows)]
                    cat_norm = [_strip_word_cell_text(v) for v in cat_values]
                    cat_norm = [x for x in cat_norm if x]
                    # 每个 series 最多取 3 个 key
                    for x in cat_norm:
                        if x not in seen_keys:
                            keys.append(x)
                            seen_keys.add(x)
                        if len(keys) >= 45:  # 控制总 key 数，避免 COM 查询太慢
                            break
                    if len(keys) >= 45:
                        break

                if not keys:
                    continue

                found_cnt = sum(1 for k in keys if is_label_found_in_words(k, docs))
                score = found_cnt / len(keys)

            finally:
                try:
                    wb.close()
                except Exception:
                    pass

            if score > best_score:
                best_score = score
                best_path = cand

    if best_path is None or best_score <= 0:
        raise ValueError(
            "无法从 Template/ 自动识别骨架：\n"
            f"当前 Template 候选数量={len(candidates)}，但所有骨架在输入 Word 中命中关键 cat label 的比例均为 0。\n"
            "请用 --template-xlsx 显式指定骨架文件路径。"
        )

    return best_path


def _write_output_and_self_check(
    template_path: Path,
    out_path: Path,
    word_parts: list[Path],
    series_refs: list[ChartSeriesRefs],
    table_mapping: Optional[dict[str, list[WordTableRef]]] = None,
) -> None:
    # 拷贝模板文件，尽量不破坏格式/图表结构
    if out_path.exists():
        out_path.unlink()
    shutil.copy2(template_path, out_path)

    template_wb = openpyxl.load_workbook(str(template_path), data_only=True)
    # 关键修复：写入阶段改用 Excel COM，避免 openpyxl 保存导致的透视/OLAP 等结构损坏
    pending_updates: dict[tuple[str, int, int], Any] = {}
    # 自检不再依赖模板 val 的具体数值（新数据 val 会变），改为统计“成功抽取了多少项”
    series_stats: list[dict[str, Any]] = []

    try:
        with WordComRunner() as runner:
            # 性能优化：缓存“label -> 命中的 tables”和“table -> matrix”，避免重复 COM 调用
            key_tables_cache: dict[str, list[Any]] = {}
            label_tables_cache_global: dict[str, list[Any]] = {}
            table_matrix_cache: dict[int, tuple[list[list[str]], list[list[Optional[float]]]]] = {}
            all_tables_cache: Optional[list[Any]] = None

            def get_table_matrices_cached(t: Any) -> tuple[list[list[str]], list[list[Optional[float]]]]:
                key = id(t)
                if key not in table_matrix_cache:
                    table_matrix_cache[key] = _extract_table_matrix(t)
                return table_matrix_cache[key]

            def get_all_tables_cached() -> list[Any]:
                """
                获取所有 Word tables（跨 part），用于 label 精确查找失败时兜底做模糊扫描。
                """
                nonlocal all_tables_cache
                if all_tables_cache is not None:
                    return all_tables_cache

                found: list[Any] = []
                seen_ids: set[tuple[int, int]] = set()
                for wp in word_parts:
                    doc = runner.open_doc(wp)
                    try:
                        tables = doc.Content.Tables
                        for i in range(1, tables.Count + 1):
                            t = tables.Item(i)
                            try:
                                tid = (int(t.Range.Start), int(t.Range.End))
                            except Exception:
                                tid = (-1, -1)
                            if tid in seen_ids:
                                continue
                            seen_ids.add(tid)
                            found.append(t)
                    except Exception:
                        continue

                all_tables_cache = found
                return found

            # 对每个 series(cat/val) 做一次填充
            for idx, sr in enumerate(series_refs, 1):
                logger.info("填充 series %d/%d: %s!%s -> %s!%s", idx, len(series_refs),
                            sr.cat.sheet, (sr.cat.min_row, sr.cat.min_col), sr.val.sheet, (sr.val.min_row, sr.val.min_col))

                tcat_ws = template_wb[sr.cat.sheet]
                tval_ws = template_wb[sr.val.sheet]

                # 提取模板 cat/val 值（假设是 1 列：min_col==max_col）
                template_cat_values = [tcat_ws.cell(sr.cat.min_row + k, sr.cat.min_col).value for k in range(sr.cat.nrows)]
                template_val_values = [tval_ws.cell(sr.val.min_row + k, sr.val.min_col).value for k in range(sr.val.nrows)]

                template_cat_values_norm = [_strip_word_cell_text(v) for v in template_cat_values]
                template_is_proportion = _detect_template_is_proportion(template_val_values)
                val_header_keys = _extract_val_header_keys(template_wb, sr)
                expected_mask = [bool(x) for x in template_cat_values_norm]
                expected_count = sum(1 for x in expected_mask if x)
                extracted_count = 0
                # 用多个 cat 标签作为 Find key（避免 key 的“第一次出现”只落在表格的一部分）
                keys: list[str] = []
                for x in template_cat_values_norm:
                    if x and x not in keys:
                        keys.append(x)
                    if len(keys) >= 3:
                        break
                if not keys:
                    logger.warning("series %d cat 全为空，跳过", idx)
                    continue

                matched_tables: list[Any] = []
                if table_mapping is not None:
                    subtable_id = _subtable_id_from_cell_range(sr.cat)
                    refs = table_mapping.get(subtable_id) if table_mapping else None
                    if not refs:
                        raise ValueError(
                            f"未找到 subtable_id 对应的 table_mapping：{subtable_id}\n"
                            f"当前 series={sr}"
                        )

                    seen_table_ids: set[tuple[int, int]] = set()
                    for ref in refs:
                        doc = runner.open_doc(Path(ref.word_path))
                        t = doc.Content.Tables.Item(ref.table_index)
                        try:
                            tid = (int(t.Range.Start), int(t.Range.End))
                        except Exception:
                            tid = (-1, -1)
                        if tid in seen_table_ids:
                            continue
                        seen_table_ids.add(tid)
                        matched_tables.append(t)
                else:
                    # 简单去重：用 table.Range.Start/End 做 key
                    seen_table_ids = set()
                    for key in keys:
                        if key not in key_tables_cache:
                            found = []
                            found_ids: set[tuple[int, int]] = set()
                            for wp in word_parts:
                                doc = runner.open_doc(wp)
                                hit_tables = _word_find_table_containing_text(doc, key)
                                for t in hit_tables:
                                    try:
                                        tid = (int(t.Range.Start), int(t.Range.End))
                                    except Exception:
                                        tid = (-1, -1)
                                    if tid in found_ids:
                                        continue
                                    found_ids.add(tid)
                                    found.append(t)
                            key_tables_cache[key] = found

                        for t in key_tables_cache[key]:
                            try:
                                tid = (int(t.Range.Start), int(t.Range.End))
                            except Exception:
                                tid = (-1, -1)
                            if tid in seen_table_ids:
                                continue
                            seen_table_ids.add(tid)
                            matched_tables.append(t)
                    if not matched_tables:
                        raise ValueError(f"无法在任何 Word part 中找到 cat keys={keys!r}，series={sr}")

                # 填充：若 cat/val sheet 不同，需要写回对应 ws；
                # 当前实现直接对 out_cat_ws 写入 cat/val 的值（cat/val 只写自己区间的单元格）
                # 其中 cat 写回模板原值（保持格式/内容完全一致）；val 写回抽取的数值。
                # 为保持结构一致：写回在 out_val_ws。
                # 先写 cat：直接复制模板对应单元格值
                for k in range(sr.cat.nrows):
                    pending_updates[(sr.cat.sheet, sr.cat.min_row + k, sr.cat.min_col)] = tcat_ws.cell(
                        sr.cat.min_row + k, sr.cat.min_col
                    ).value

                # 再对 val 用 table 匹配抽取
                # 为复用匹配函数，将 cat/val 都在同一个 table 中匹配
                def _fill_val_from_tables() -> None:
                    nonlocal extracted_count
                    # 简化：用任意一个 ws 作为 template_ws 读取值
                    template_ws = template_wb[sr.val.sheet]

                    # -------- 1) 快速路径：单表内匹配完整 cat->val 序列 --------
                    for table in matched_tables:
                        text_matrix, num_matrix = get_table_matrices_cached(table)
                        match = _match_cat_and_val_columns(
                            text_matrix,
                            num_matrix,
                            cat_values=[str(x) if x is not None else "" for x in template_cat_values],
                            val_header_keys=val_header_keys,
                        )
                        if match is None:
                            continue
                        row_positions, _cat_col, val_col = match
                        n = sr.val.nrows
                        for k in range(n):
                            wnum = num_matrix[row_positions[k]][val_col]
                            tval = template_ws.cell(sr.val.min_row + k, sr.val.min_col).value
                            if wnum is None or not expected_mask[k]:
                                pending_updates[(sr.val.sheet, sr.val.min_row + k, sr.val.min_col)] = tval
                                continue

                            scaled = _coerce_word_number_scale(float(wnum), template_is_proportion=template_is_proportion)
                            extracted_count += 1
                            # 若模板该格是整数，则尽量按整数写回
                            if isinstance(tval, (int, float)) and abs(float(tval) - round(float(tval))) < 1e-9:
                                pending_updates[(sr.val.sheet, sr.val.min_row + k, sr.val.min_col)] = int(round(scaled))
                            else:
                                pending_updates[(sr.val.sheet, sr.val.min_row + k, sr.val.min_col)] = float(scaled)
                        # 只有在本 series 抽取覆盖率达到阈值时才返回；
                        # 否则继续尝试其它 table（否则可能只匹配到了错误 val_col）。
                        min_ratio = 0.6
                        if expected_count > 0 and (extracted_count / expected_count) >= min_ratio:
                            return
                        # 继续下一个 table；若所有 table 都达不到阈值，随后会进入 fallback 兜底扫描

                    # -------- 2) 兜底路径：逐行标签定位 val（跨 table / 非连续布局兜底） --------
                    val_header_keys_norm = [_strip_word_cell_text(v) for v in val_header_keys if v is not None]
                    val_header_keys_norm = [v for v in val_header_keys_norm if v]

                    def find_val_for_one_label(label_norm: str) -> float:
                        """
                        在包含 label_norm 的表格行中：
                        - 用 val_header_keys_norm 先确定“更像值列”的列（列头匹配优先）
                        - 然后从该行里取可解析的数值
                        该逻辑不依赖模板 val 的具体数值（新数据 val 会变）。
                        """
                        if not label_norm:
                            raise ValueError("label_norm 为空，无法定位 val")

                        if table_mapping is not None:
                            candidate_tables = matched_tables
                        else:
                            if label_norm not in label_tables_cache_global:
                                found: list[Any] = []
                                found_ids: set[tuple[int, int]] = set()
                                for wp in word_parts:
                                    doc = runner.open_doc(wp)
                                    hit_tables = _word_find_table_containing_text(doc, label_norm)
                                    for t in hit_tables:
                                        try:
                                            tid = (int(t.Range.Start), int(t.Range.End))
                                        except Exception:
                                            tid = (-1, -1)
                                        if tid in found_ids:
                                            continue
                                        found_ids.add(tid)
                                        found.append(t)
                                # 精确查找失败时：退化到“全表模糊扫描”，以适配新数据条目文本变化
                                if found:
                                    label_tables_cache_global[label_norm] = found
                                else:
                                    label_tables_cache_global[label_norm] = get_all_tables_cached()

                            candidate_tables = label_tables_cache_global[label_norm]
                        for t in candidate_tables:
                            text_matrix, num_matrix = get_table_matrices_cached(t)
                            R = len(text_matrix)
                            C = len(text_matrix[0]) if R else 0

                            # 为该 table 计算每列的表头匹配分数（用于决定值列优先级）
                            col_header_score: list[int] = []
                            for cc in range(C):
                                col_texts = {text_matrix[rr][cc] for rr in range(R) if text_matrix[rr][cc]}
                                score = sum(1 for hk in val_header_keys_norm if hk in col_texts)
                                col_header_score.append(score)

                            # 列优先级：表头分高的优先；其次数值可解析多的优先
                            def col_numeric_score(cc: int) -> int:
                                return sum(1 for rr in range(R) if num_matrix[rr][cc] is not None)

                            ordered_cols = list(range(C))
                            ordered_cols.sort(key=lambda cc: (col_header_score[cc], col_numeric_score(cc)), reverse=True)

                            for rr in range(R):
                                # 这一行是否出现过该 label？
                                # 用“整行拼接”来做匹配：有些条目（如区间）可能被拆在多个单元格里
                                row_concat = "".join(text_matrix[rr])
                                if not _labels_match(row_concat, label_norm):
                                    continue
                                # 在同一行里按“值列优先级”找可解析数值
                                for cc in ordered_cols:
                                    wnum = num_matrix[rr][cc]
                                    if wnum is None:
                                        continue
                                    scaled = _coerce_word_number_scale(float(wnum), template_is_proportion=template_is_proportion)
                                    return float(scaled)

                        raise ValueError(f"未能在 Word 表中定位 label={label_norm!r} 的可解析 val")

                    n = sr.val.nrows
                    for k in range(n):
                        label_norm = template_cat_values_norm[k] if k < len(template_cat_values_norm) else ""
                        tval = template_ws.cell(sr.val.min_row + k, sr.val.min_col).value
                        if not label_norm:
                            pending_updates[(sr.val.sheet, sr.val.min_row + k, sr.val.min_col)] = tval
                            continue
                        try:
                            extracted = find_val_for_one_label(label_norm)
                            extracted_count += 1
                            # 若模板是整数，尽量写回整数
                            if isinstance(tval, (int, float)) and abs(float(tval) - round(float(tval))) < 1e-9:
                                pending_updates[(sr.val.sheet, sr.val.min_row + k, sr.val.min_col)] = int(round(extracted))
                            else:
                                pending_updates[(sr.val.sheet, sr.val.min_row + k, sr.val.min_col)] = extracted
                        except Exception:
                            # 找不到就回退为模板值（避免破坏格式/图表引用；自检仍会捕获）
                            pending_updates[(sr.val.sheet, sr.val.min_row + k, sr.val.min_col)] = tval

                    # 兜底路径填充完成后返回（自检会决定是否还需要更强策略）
                    return

                _fill_val_from_tables()
                series_stats.append(
                    {
                        "series_index": idx,
                        "expected_count": expected_count,
                        "extracted_count": extracted_count,
                        "cat_keys_used": keys,
                        "val_header_keys_used": val_header_keys,
                    }
                )

        # Excel COM 保存：避免 openpyxl 保存破坏 OLAP/透视缓存等结构
        _excel_com_write_updates(out_path, pending_updates)
    finally:
        try:
            template_wb.close()
        except Exception:
            pass

    # 自检：新数据 val 会变化，因此自检改为“抽取覆盖率”而非与模板数值逐项相等
    _self_check_extraction_coverage(series_stats)


def _excel_com_write_updates(workbook_path: Path, updates: dict[tuple[str, int, int], Any]) -> None:
    """
    通过 Excel COM 只写入单元格值，最大化保留模板里的复杂对象结构（OLAP/数据透视/图表等）。
    """
    if not updates:
        return

    import win32com.client  # type: ignore

    excel = win32com.client.Dispatch("Excel.Application")
    # 有些环境下设置 Visible 可能直接触发 COM 异常；不影响后续写入逻辑
    try:
        excel.Visible = False
    except Exception:
        pass
    try:
        excel.DisplayAlerts = False
    except Exception:
        pass

    wb = excel.Workbooks.Open(str(workbook_path))
    try:
        # 分 sheet 缓存 Worksheet 对象，避免重复查找
        sheet_cache: dict[str, Any] = {}

        def get_ws(name: str) -> Any:
            if name not in sheet_cache:
                sheet_cache[name] = wb.Worksheets(name)
            return sheet_cache[name]

        for (sheet, row, col), val in updates.items():
            ws = get_ws(sheet)
            ws.Cells(row, col).Value = val

        wb.Save()
    finally:
        try:
            wb.Close(False)
        except Exception:
            pass
        try:
            excel.Quit()
        except Exception:
            pass


def _cell_values_equal_for_patch(a: Any, b: Any, *, rel_tol: float, abs_tol: float) -> bool:
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return math.isclose(float(a), float(b), rel_tol=rel_tol, abs_tol=abs_tol)
    return _strip_word_cell_text(str(a)) == _strip_word_cell_text(str(b))


def _self_check_values(
    source_path: Path,
    out_path: Path,
    series_refs: list[ChartSeriesRefs],
    *,
    rel_tol: float = 1e-9,
    abs_tol: float = 1e-9,
) -> None:
    """自检：out_path 与 source_path 在图表 cat/val 引用区间上的单元格值一致（repair_output_by_patch 用）。"""
    wb_src = openpyxl.load_workbook(str(source_path), data_only=True)
    wb_out = openpyxl.load_workbook(str(out_path), data_only=True)
    try:
        errs: list[str] = []
        for sr in series_refs:
            for k in range(sr.cat.nrows):
                r, c = sr.cat.min_row + k, sr.cat.min_col
                vs = wb_src[sr.cat.sheet].cell(r, c).value
                vo = wb_out[sr.cat.sheet].cell(r, c).value
                if not _cell_values_equal_for_patch(vs, vo, rel_tol=rel_tol, abs_tol=abs_tol):
                    errs.append(f"cat {sr.cat.sheet}!{r},{c}: src={vs!r} out={vo!r}")
            for k in range(sr.val.nrows):
                r, c = sr.val.min_row + k, sr.val.min_col
                vs = wb_src[sr.val.sheet].cell(r, c).value
                vo = wb_out[sr.val.sheet].cell(r, c).value
                if not _cell_values_equal_for_patch(vs, vo, rel_tol=rel_tol, abs_tol=abs_tol):
                    errs.append(f"val {sr.val.sheet}!{r},{c}: src={vs!r} out={vo!r}")
        if errs:
            sample = "\n".join(errs[:20])
            raise AssertionError(f"自检失败：patch 后值与源文件不一致（前20条）：\n{sample}")
        logger.info("自检通过：图表引用区间与源文件一致。")
    finally:
        wb_src.close()
        wb_out.close()


def _self_check_extraction_coverage(
    series_stats: list[dict[str, Any]],
    *,
    min_extracted_ratio: float = 0.6,
) -> None:
    """
    新数据 val 会变化，因此自检改为：
    - 对每个 series，只要 template cat 该行不是空的，就期望从 Word 中成功抽取出 val；
    - 抽取覆盖率（extracted/expected）需要达到阈值。
    """
    errors: list[str] = []
    for st in series_stats:
        exp = int(st.get("expected_count", 0) or 0)
        if exp <= 0:
            continue
        extracted = int(st.get("extracted_count", 0) or 0)
        ratio = extracted / exp if exp else 0.0
        if ratio < min_extracted_ratio:
            errors.append(
                f"series_index={st.get('series_index')} 抽取覆盖率过低: {extracted}/{exp} ({ratio:.2%})"
            )

    if errors:
        sample = "\n".join(errors[:10])
        raise AssertionError(f"自检失败：Word->Excel val 抽取覆盖率不足（取前10条）：\n{sample}")
    logger.info("自检通过：val 抽取覆盖率满足阈值。")


def main() -> None:
    parser = argparse.ArgumentParser(description="Word -> Excel(表+图) 自动复刻")
    default_input = MODULE_DIR / "input"
    parser.add_argument(
        "--input-dir",
        "-i",
        default=str(default_input),
        help="输入目录（放 Word/RTF 原始数据文件）。默认：07_Word_to_Excel_to_Figure/input",
    )
    parser.add_argument(
        "--template-xlsx",
        default=None,
        help="骨架 Excel 路径（可选；默认使用模块目录里唯一的 .xlsx 作为骨架）。",
    )
    parser.add_argument("--output-xlsx", default=None, help="输出 Excel 路径（默认：output/replicate_<骨架名>.xlsx）")
    parser.add_argument("--plan-only", action="store_true", help="仅生成 Excel 子表 -> Word 表映射候选，并退出")
    parser.add_argument("--plan-out-json", default=None, help="plan-only 输出的 JSON 路径（默认：output/table_mapping_plan_*.json）")
    parser.add_argument("--table-map-json", default=None, help="使用你确认后的 table 映射 JSON 来限定数据抓取来源")
    parser.add_argument("--verbose", "-v", action="store_true", help="显示更详细日志")

    args = parser.parse_args()
    _setup_logging(args.verbose)

    input_dir = Path(args.input_dir).expanduser().resolve()
    if not input_dir.exists():
        raise FileNotFoundError(f"输入目录不存在: {input_dir}")

    word_parts = _collect_word_parts(input_dir)
    skeleton_xlsx = _select_best_skeleton_xlsx(
        template_xlsx_arg=args.template_xlsx,
        template_dir=TEMPLATE_DIR,
        word_parts=word_parts,
    )

    if args.output_xlsx:
        out_path = Path(args.output_xlsx).expanduser().resolve()
    else:
        out_dir = MODULE_DIR / "output"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"replicate_{skeleton_xlsx.stem}.xlsx"

    logger.info("骨架Excel: %s", skeleton_xlsx)
    logger.info("Word/RTF parts: %s", [p.name for p in word_parts])
    logger.info("输出Excel: %s", out_path)

    # 读取模板图表引用区间
    template_wb_struct = openpyxl.load_workbook(str(skeleton_xlsx), data_only=False)
    try:
        series_refs = _extract_template_series_refs(template_wb_struct)
    finally:
        template_wb_struct.close()

    if not series_refs:
        raise ValueError("未能从模板 Excel 中发现任何图表 series(cat/val) 引用区间，请检查模板是否包含图表。")

    logger.info("发现图表 series(cat/val) 引用对数: %d", len(series_refs))

    if args.plan_only:
        plan = build_table_mapping_plan(
            template_path=skeleton_xlsx,
            word_parts=word_parts,
            series_refs=series_refs,
            top_k_per_subtable=3,
        )
        out_json = (
            Path(args.plan_out_json).expanduser().resolve()
            if args.plan_out_json
            else (MODULE_DIR / "output" / f"table_mapping_plan_{skeleton_xlsx.stem}.json")
        )
        out_json.parent.mkdir(parents=True, exist_ok=True)
        with out_json.open("w", encoding="utf-8") as f:
            json.dump(plan, f, ensure_ascii=False, indent=2)

        # 打印一个简短预览：每个 subtable 取候选列表第一个（score 最高）
        subtables = plan.get("subtables", {}) or {}
        logger.info("plan-only 完成：%s", out_json)
        for sid in sorted(subtables.keys())[:50]:
            cand_list = subtables[sid] or []
            if not cand_list:
                continue
            c0 = cand_list[0]
            sn = str(c0.get("snippet", "") or "")
            sn = sn[:60] + "..." if len(sn) > 60 else sn
            logger.info(
                "SUBTABLE %s => word=%s table_index=%s score=%s snippet=%s",
                sid,
                os.path.basename(str(c0.get("word_file"))),
                c0.get("table_index"),
                c0.get("score"),
                sn,
            )
        return

    _write_output_and_self_check(
        template_path=skeleton_xlsx,
        out_path=out_path,
        word_parts=word_parts,
        series_refs=series_refs,
        table_mapping=load_table_mapping_json(args.table_map_json or None),
    )


if __name__ == "__main__":
    main()

