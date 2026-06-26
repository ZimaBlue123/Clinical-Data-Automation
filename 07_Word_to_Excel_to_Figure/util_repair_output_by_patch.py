"""
快速修复版本：把“已经算对了图表数据的 output”里的 chart.series(cat/val) 区间值
patch 回“模板 Excel”副本上，用 Excel COM 写入以保留 OLAP/数据透视结构。

用途：
- 修复微软 Office 打开时的“部分内容有问题/需要修复”弹窗（通常是 openpyxl 保存破坏了复杂对象）。
- 避免重新跑 Word->数据->图表的重耗时提取步骤。
"""

from __future__ import annotations

import argparse
import shutil
from pathlib import Path
from typing import Any

import openpyxl

import word_to_excel_to_figure as m


def _load_values(wb: openpyxl.Workbook, sheet_name: str, row: int, col: int) -> Any:
    ws = wb[sheet_name]
    return ws.cell(row, col).value


def main() -> None:
    parser = argparse.ArgumentParser(description="修复 output 文件结构（保留模板的复杂对象）")
    parser.add_argument("--template-xlsx", required=True, help="模板 Excel 路径")
    parser.add_argument("--source-xlsx", required=True, help="旧 output（数据已正确）路径")
    parser.add_argument("--output-xlsx", required=True, help="修复后成品输出路径")
    parser.add_argument("--verbose", "-v", action="store_true", help="显示更多日志")
    args = parser.parse_args()

    template_path = Path(args.template_xlsx).expanduser().resolve()
    source_path = Path(args.source_xlsx).expanduser().resolve()
    out_path = Path(args.output_xlsx).expanduser().resolve()

    if not template_path.is_file():
        raise FileNotFoundError(f"模板不存在: {template_path}")
    if not source_path.is_file():
        raise FileNotFoundError(f"source output 不存在: {source_path}")

    # 只为读值：data_only=True
    twb_val = openpyxl.load_workbook(str(template_path), data_only=True)
    swb_val = openpyxl.load_workbook(str(source_path), data_only=True)
    twb_struct = openpyxl.load_workbook(str(template_path), data_only=False)

    try:
        series_refs = m._extract_template_series_refs(twb_struct)
    finally:
        twb_val.close()
        swb_val.close()
        twb_struct.close()

    if out_path.exists():
        out_path.unlink()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, out_path)

    # 构建 updates：仅 chart.series cat/val 引用区间的 cell values
    updates: dict[tuple[str, int, int], Any] = {}
    twb_val = openpyxl.load_workbook(str(source_path), data_only=True)
    try:
        for sr in series_refs:
            n = sr.cat.nrows
            for k in range(n):
                updates[(sr.cat.sheet, sr.cat.min_row + k, sr.cat.min_col)] = _load_values(
                    twb_val, sr.cat.sheet, sr.cat.min_row + k, sr.cat.min_col
                )
            n = sr.val.nrows
            for k in range(n):
                updates[(sr.val.sheet, sr.val.min_row + k, sr.val.min_col)] = _load_values(
                    twb_val, sr.val.sheet, sr.val.min_row + k, sr.val.min_col
                )
    finally:
        twb_val.close()

    # Excel COM 写入：修复复杂对象结构
    m._excel_com_write_updates(out_path, updates)

    # 自检：patch 后图表 cat/val 区间应与 source（数据已正确的旧 output）一致
    m._self_check_values(source_path, out_path, series_refs)


if __name__ == "__main__":
    # 继承主模块日志设置
    m._setup_logging(verbose=False)
    main()

