from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.serology_utils import OUTPUT_MARKERS, canonical_sample_id


def _cell_to_str(v: Any) -> str:
    if v is None:
        return ""
    # 保留原值语义，但统一成字符串并去掉首尾空白
    return str(v).strip()


def load_serology_excel(path: Path) -> dict[str, dict[str, tuple[str, str]]]:
    """
    读取血清汇总 Excel（固定结构）：
    - A 列：样品 ID
    - 每个指标占 2 列：value / note
    - 第一张 sheet 为主数据
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    out: dict[str, dict[str, tuple[str, str]]] = {}
    # 数据从第 3 行开始（README/模块实现一致）
    for r in range(3, ws.max_row + 1):
        sid_raw = ws.cell(r, 1).value
        sid = canonical_sample_id(_cell_to_str(sid_raw))
        if not sid:
            continue

        marker_map: dict[str, tuple[str, str]] = {}
        for i, m in enumerate(OUTPUT_MARKERS):
            # value 列：2,4,6,8,10；note 列：3,5,7,9,11
            v_col = 2 + i * 2
            n_col = v_col + 1
            v = _cell_to_str(ws.cell(r, v_col).value)
            n = _cell_to_str(ws.cell(r, n_col).value)
            marker_map[m] = (v, n)

        out[sid] = marker_map

    wb.close()
    return out


def main() -> None:
    parser = argparse.ArgumentParser(description="对比 PDF 输出 Excel 与 Word 汇总 Excel（血清五项）")
    parser.add_argument("--pdf-excel", type=Path, required=True, help="10_PDF_Batch_to_Excel 输出的 serology_report_merged*.xlsx")
    parser.add_argument("--word-excel", type=Path, required=True, help="09_Word_All_Tables_to_Excel 输出的 word_tables_merged.xlsx")
    parser.add_argument("--out-csv", type=Path, default=None, help="可选：将差异明细写入 CSV")
    args = parser.parse_args()

    pdf_path = args.pdf_excel
    word_path = args.word_excel
    if not pdf_path.exists():
        raise FileNotFoundError(f"pdf-excel 不存在：{pdf_path}")
    if not word_path.exists():
        raise FileNotFoundError(f"word-excel 不存在：{word_path}")

    pdf = load_serology_excel(pdf_path)
    word = load_serology_excel(word_path)

    pdf_ids = set(pdf.keys())
    word_ids = set(word.keys())
    common_ids = pdf_ids & word_ids

    only_pdf = sorted(pdf_ids - word_ids)
    only_word = sorted(word_ids - pdf_ids)

    missing_in_pdf = 0
    missing_in_word = 0
    value_or_note_diff = 0
    discrepancies: list[dict[str, str]] = []

    for sid in sorted(common_ids):
        pdf_map = pdf[sid]
        word_map = word[sid]

        for m in OUTPUT_MARKERS:
            pv, pn = pdf_map.get(m, ("", ""))
            wv, wn = word_map.get(m, ("", ""))

            pdf_empty = not (pv or pn)
            word_empty = not (wv or wn)

            if word_empty and not pdf_empty:
                # Word 端空，PDF 有值：这不是“PDF 缺项”，只统计值差异即可
                pass
            if pdf_empty and not word_empty:
                missing_in_pdf += 1
            if word_empty and not pdf_empty:
                missing_in_word += 1

            if (pv, pn) != (wv, wn):
                value_or_note_diff += 1
                discrepancies.append(
                    {
                        "sample_id": sid,
                        "marker": m,
                        "pdf_value": pv,
                        "pdf_note": pn,
                        "word_value": wv,
                        "word_note": wn,
                    }
                )

    print("=== 血清报告对比（PDF vs Word）===")
    print(f"PDF 样品数：{len(pdf_ids)}")
    print(f"Word 样品数：{len(word_ids)}")
    print(f"交集样品数：{len(common_ids)}")
    print(f"仅 PDF：{len(only_pdf)}（示例：{only_pdf[:10]}）")
    print(f"仅 Word：{len(only_word)}（示例：{only_word[:10]}）")
    print(f"PDF 缺项（Word 有值但 PDF 空）：{missing_in_pdf}")
    print(f"Word 缺项（PDF 有值但 Word 空）：{missing_in_word}")
    print(f"值/说明不同的指标块：{value_or_note_diff}")

    if args.out_csv:
        args.out_csv.parent.mkdir(parents=True, exist_ok=True)
        with args.out_csv.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=[
                    "sample_id",
                    "marker",
                    "pdf_value",
                    "pdf_note",
                    "word_value",
                    "word_note",
                ],
            )
            writer.writeheader()
            for row in discrepancies:
                writer.writerow(row)
        print(f"差异明细已写入：{args.out_csv}")


if __name__ == "__main__":
    main()

